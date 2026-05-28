import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Toplevel
import pandas as pd
import threading
import os
import sys
import subprocess
import time
import re
import csv
from html import unescape
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Dictionary file paths - CSV format
RULES_FILE_CSV = 'translation_rules.csv'
IGNORE_FILE_CSV = 'ignore_rules.csv'
BLACKLIST_FILE_CSV = 'blacklist_terms.csv'

# Legacy TXT file paths (for backward compatibility)
RULES_FILE = 'translation_rules.txt'
IGNORE_FILE = 'ignore_rules.txt'
BLACKLIST_FILE = 'blacklist_terms.txt'
APP_VERSION = "v18"  # Restored processing mode section; blacklist action as radio buttons (Видаляти/Повідомляти)
MIN_SHORT_DESC_WORDS = 2
MIN_SHORT_DESC_LENGTH = 24
DECIMAL_MAX_INT_DIGITS = 4
DECIMAL_MAX_FRACTION_DIGITS = 2
DECIMAL_CONTEXT_UNITS = r'%|°|кг|г|мг|л|мл|см|мм|м|дюйм|kg|g|mg|l|ml|cm|mm|m|inch|in|x|×'

DEFAULT_RULES = """# Формат: російський_корінь = українські_корені_через_кому
# Наприклад:
питани = харчуванн
лук = цибул
настройк = настройк, настроюванн
вид = вигляд
пользовател = користувач
"""

DEFAULT_IGNORES = """# Додайте сюди фрази (українською), які є ПРАВИЛЬНИМИ винятками.
# Якщо програма знайде цю фразу у реченні, вона проігнорує помилку.
# Кожна фраза з нового рядка:
збалансоване харчування
дитяче харчування
здорове харчування
зелена цибуля
ріпчаста цибуля
"""

DEFAULT_BLACKLIST = """# Додайте сюди слова та фрази (будь-якою мовою), які є ПІДОЗРІЛИМИ або ЗАБОРОНЕНІ.
# Якщо програма знайде цю фразу в українському описі, вона позначить рядок як помилковий.
# Також програма автоматично видалить речення або пункт списку, де знайдено заборонений термін.
# Кожна фраза з нового рядка:
www.
http://
@gmail.
e-mail:
копія
репліка
"""

# CSV default headers and sample data
DEFAULT_RULES_CSV_HEADER = ['російський_корінь', 'українські_корені', 'виключення', 'тип', 'коментар']
DEFAULT_IGNORES_CSV_HEADER = ['фраза', 'категорія', 'коментар']
DEFAULT_BLACKLIST_CSV_HEADER = ['термін', 'виключення', 'дія', 'категорія', 'коментар']


def migrate_txt_to_csv():
    """Migrate old TXT dictionaries to CSV format if they exist."""
    
    # Migrate translation rules
    if os.path.exists(RULES_FILE) and not os.path.exists(RULES_FILE_CSV):
        print(f"Міграція {RULES_FILE} → {RULES_FILE_CSV}...")
        rules_data = []
        with open(RULES_FILE, 'r', encoding='utf-8') as f:
            current_comment = ''
            for line in f:
                line = line.rstrip('\n')
                if line.strip().startswith('#'):
                    # Extract comment
                    current_comment = line.strip('#').strip()
                elif '=' in line:
                    # Parse rule
                    parts = line.split('#', 1)
                    rule_part = parts[0].strip()
                    inline_comment = parts[1].strip() if len(parts) > 1 else ''
                    
                    if '=' in rule_part:
                        ru, ua_list = rule_part.split('=', 1)
                        ru_stem = ru.strip()
                        ua_stems = ua_list.strip()
                        comment = inline_comment or current_comment
                        
                        rules_data.append({
                            'російський_корінь': ru_stem,
                            'українські_корені': ua_stems,
                            'виключення': '',
                            'тип': '',
                            'коментар': comment
                        })
                        current_comment = ''  # Reset after use
        
        # Write CSV
        with open(RULES_FILE_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=DEFAULT_RULES_CSV_HEADER)
            writer.writeheader()
            writer.writerows(rules_data)
        print(f"✓ Міграцію завершено: {len(rules_data)} правил")
    
    # Migrate ignore rules
    if os.path.exists(IGNORE_FILE) and not os.path.exists(IGNORE_FILE_CSV):
        print(f"Міграція {IGNORE_FILE} → {IGNORE_FILE_CSV}...")
        ignores_data = []
        with open(IGNORE_FILE, 'r', encoding='utf-8') as f:
            current_comment = ''
            for line in f:
                line = line.rstrip('\n')
                if line.strip().startswith('#'):
                    current_comment = line.strip('#').strip()
                elif line.strip():
                    parts = line.split('#', 1)
                    phrase = parts[0].strip()
                    inline_comment = parts[1].strip() if len(parts) > 1 else ''
                    comment = inline_comment or current_comment
                    
                    ignores_data.append({
                        'фраза': phrase,
                        'категорія': '',
                        'коментар': comment
                    })
                    current_comment = ''
        
        with open(IGNORE_FILE_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=DEFAULT_IGNORES_CSV_HEADER)
            writer.writeheader()
            writer.writerows(ignores_data)
        print(f"✓ Міграцію завершено: {len(ignores_data)} фраз")
    
    # Migrate blacklist
    if os.path.exists(BLACKLIST_FILE) and not os.path.exists(BLACKLIST_FILE_CSV):
        print(f"Міграція {BLACKLIST_FILE} → {BLACKLIST_FILE_CSV}...")
        blacklist_data = []
        with open(BLACKLIST_FILE, 'r', encoding='utf-8') as f:
            current_comment = ''
            for line in f:
                line = line.rstrip('\n')
                if line.strip().startswith('#'):
                    current_comment = line.strip('#').strip()
                elif line.strip():
                    parts = line.split('#', 1)
                    term = parts[0].strip()
                    inline_comment = parts[1].strip() if len(parts) > 1 else ''
                    comment = inline_comment or current_comment
                    
                    blacklist_data.append({
                        'термін': term,
                        'виключення': '',
                        'дія': 'Видалити',  # За замовчуванням - як зараз працює
                        'категорія': '',
                        'коментар': comment
                    })
                    current_comment = ''
        
        with open(BLACKLIST_FILE_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=DEFAULT_BLACKLIST_CSV_HEADER)
            writer.writeheader()
            writer.writerows(blacklist_data)
        print(f"✓ Міграцію завершено: {len(blacklist_data)} термінів")


def create_default_csv_files():
    """Create default CSV files if they don't exist."""
    
    # Create default translation_rules.csv
    if not os.path.exists(RULES_FILE_CSV):
        with open(RULES_FILE_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=DEFAULT_RULES_CSV_HEADER)
            writer.writeheader()
            # Add sample data
            writer.writerow({
                'російський_корінь': 'питани',
                'українські_корені': 'харчуванн',
                'виключення': 'збалансоване харчування,дитяче харчування,здорове харчування',
                'тип': 'Помилка перекладу',
                'коментар': 'Питание (електричне) → живлення, а не харчування'
            })
    
    # Create default ignore_rules.csv
    if not os.path.exists(IGNORE_FILE_CSV):
        with open(IGNORE_FILE_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=DEFAULT_IGNORES_CSV_HEADER)
            writer.writeheader()
            writer.writerow({
                'фраза': 'збалансоване харчування',
                'категорія': 'Харчування',
                'коментар': 'Допустимий контекст для слова "харчування"'
            })
    
    # Create default blacklist_terms.csv
    if not os.path.exists(BLACKLIST_FILE_CSV):
        with open(BLACKLIST_FILE_CSV, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=DEFAULT_BLACKLIST_CSV_HEADER)
            writer.writeheader()
            writer.writerow({
                'термін': 'www.',
                'категорія': 'URL',
                'коментар': 'Заборонені веб-адреси в описах'
            })


UKRAINIAN_MARKERS = {
    'і', 'та', 'це', 'для', 'що', 'від', 'під', 'при', 'через', 'згідно', 'наявність',
    'зручність', 'технічні', 'обслуговування', 'попередження', 'увага', 'примітка', 'модель'
}
RUSSIAN_MARKERS = {
    'и', 'или', 'это', 'для', 'что', 'при', 'через', 'согласно', 'наличие',
    'удобство', 'технические', 'обслуживание', 'внимание', 'примечание', 'модель'
}
UKRAINIAN_UNIQUE_CHARS = set('іїєґ')
RUSSIAN_UNIQUE_CHARS = set('ыэёъ')
CHINESE_CHAR_PATTERN = re.compile(r'[\u4e00-\u9fff]')
TECHNICAL_HTML_PATTERNS = [
    r'(?is)<style\b[^>]*>.*?</style>',
    r'(?is)<script\b[^>]*>.*?</script>',
    r'(?is)<div\b[^>]*class="[^"]*ssd-module-wrap[^"]*"[^>]*>.*?</div>',
    r'(?is)<div\b[^>]*class=["\']?ssd-module[^>]*>.*?</div>',
    r'(?is)<div\b[^>]*(cssurl|skucode|skudesign|id="zbViewModulesH"|id="zbViewModulesHeight")[^>]*>.*?</div>',
    r'(?is)<input\b[^>]*(zbViewModulesHeight)[^>]*>',
]


class TextEditor(Toplevel):
    def __init__(self, parent, title, file_path, default_content):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title(title)
        self.geometry("650x450")
        self.file_path = file_path

        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill='both', expand=True)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', side='bottom', pady=(10, 0))

        self.btn_save = ttk.Button(btn_frame, text="Зберегти", command=self.save_and_close)
        self.btn_save.pack(side='right', padx=10)

        self.btn_cancel = ttk.Button(btn_frame, text="Скасувати", command=self.destroy)
        self.btn_cancel.pack(side='right')

        self.txt_content = tk.Text(main_frame, wrap='word', font=('Consolas', 11))
        self.txt_content.pack(fill='both', expand=True, side='top')

        if not os.path.exists(file_path):
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(default_content)

        with open(file_path, 'r', encoding='utf-8') as f:
            self.txt_content.insert('1.0', f.read())

    def save_and_close(self):
        content = self.txt_content.get('1.0', tk.END).strip()
        with open(self.file_path, 'w', encoding='utf-8') as f:
            f.write(content + "\n")
        self.destroy()


class DictionaryManager(Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title("📚 Словники")
        self.geometry("900x600")
        self.parent = parent
        
        # Dictionary files configuration
        self.dictionaries = {
            'rules': {
                'file': RULES_FILE,
                'default': DEFAULT_RULES,
                'title': '🔄 Помилки перекладу'
            },
            'blacklist': {
                'file': BLACKLIST_FILE,
                'default': DEFAULT_BLACKLIST,
                'title': '⛔ Чорний список'
            }
        }
        
        self.create_widgets()
        self.load_all_dictionaries()
        self.update_statistics()
        self.bind_cyrillic_shortcuts()
    
    def bind_cyrillic_shortcuts(self):
        """Прив'язка кириличних клавіатурних скорочень"""
        self.bind('<Control-Key-с>', lambda e: self.handle_copy(e))
        self.bind('<Control-Key-С>', lambda e: self.handle_copy(e))
        self.bind('<Control-Key-м>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-М>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-в>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-В>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-ч>', lambda e: self.handle_cut(e))
        self.bind('<Control-Key-Ч>', lambda e: self.handle_cut(e))
    
    def handle_copy(self, event):
        """Обробка копіювання"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'selection_get'):
                widget.event_generate('<<Copy>>')
        except:
            pass
        return 'break'
    
    def handle_paste(self, event):
        """Обробка вставки"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'insert'):
                widget.event_generate('<<Paste>>')
        except:
            pass
        return 'break'
    
    def handle_cut(self, event):
        """Обробка вирізання"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'selection_get'):
                widget.event_generate('<<Cut>>')
        except:
            pass
        return 'break'
        
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill='both', expand=True)
        
        # Statistics frame at the top
        stats_frame = ttk.LabelFrame(main_frame, text="📊 Статистика", padding=5)
        stats_frame.pack(fill='x', pady=(0, 10))
        
        self.lbl_stats = ttk.Label(stats_frame, text="Завантаження...", font=('Segoe UI', 9))
        self.lbl_stats.pack(anchor='w', padx=5)
        
        # Search frame
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(search_frame, text="🔍 Пошук:", font=('Segoe UI', 9, 'bold')).pack(side='left', padx=5)
        self.entry_search = ttk.Entry(search_frame, width=50)
        self.entry_search.pack(side='left', padx=5, fill='x', expand=True)
        self.entry_search.bind('<KeyRelease>', self.on_search)
        
        self.btn_clear_search = ttk.Button(search_frame, text="Очистити", command=self.clear_search)
        self.btn_clear_search.pack(side='left', padx=5)
        
        # Notebook (tabs)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(0, 10))
        
        # Create tabs for each dictionary and maintain tab-to-key mapping
        self.text_widgets = {}
        self.tab_to_key = {}  # Map tab index to dictionary key
        
        for idx, (key, dict_info) in enumerate(self.dictionaries.items()):
            tab_frame = ttk.Frame(self.notebook, padding=5)
            self.notebook.add(tab_frame, text=dict_info['title'])
            self.tab_to_key[idx] = key  # Store explicit mapping
            
            # Text widget with scrollbar
            text_widget = tk.Text(tab_frame, wrap='word', font=('Consolas', 10), undo=True)
            scrollbar = ttk.Scrollbar(tab_frame, orient='vertical', command=text_widget.yview)
            text_widget.config(yscrollcommand=scrollbar.set)
            
            text_widget.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            self.text_widgets[key] = text_widget
        
        # Bottom button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x')
        
        self.btn_save_all = ttk.Button(btn_frame, text="💾 Зберегти всі", command=self.save_all)
        self.btn_save_all.pack(side='left', padx=5)
        
        self.btn_reload = ttk.Button(btn_frame, text="🔄 Перезавантажити", command=self.reload_all)
        self.btn_reload.pack(side='left', padx=5)
        
        self.btn_close = ttk.Button(btn_frame, text="Закрити", command=self.destroy)
        self.btn_close.pack(side='right', padx=5)
        
    def load_all_dictionaries(self):
        """Load content from all dictionary files"""
        for key, dict_info in self.dictionaries.items():
            file_path = dict_info['file']
            default_content = dict_info['default']
            
            # Create file with default content if it doesn't exist
            if not os.path.exists(file_path):
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(default_content)
            
            # Load content
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            text_widget = self.text_widgets[key]
            text_widget.delete('1.0', tk.END)
            text_widget.insert('1.0', content)
            
    def save_all(self):
        """Save all dictionaries to their respective files"""
        try:
            for key, dict_info in self.dictionaries.items():
                file_path = dict_info['file']
                text_widget = self.text_widgets[key]
                content = text_widget.get('1.0', tk.END).strip()
                
                # Add newline at end if content exists
                if content:
                    content += '\n'
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
            
            self.update_statistics()
            messagebox.showinfo("Успіх", "Всі словники успішно збережено!", parent=self)
            
            # Log to parent app if available
            if hasattr(self.parent, 'log_message'):
                self.parent.log_message("Словники оновлено через Менеджер словників")
                
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося зберегти словники:\n{e}", parent=self)
            
    def reload_all(self):
        """Reload all dictionaries from files"""
        response = messagebox.askyesno(
            "Підтвердження", 
            "Перезавантажити всі словники з файлів? Незбережені зміни будуть втрачені.",
            parent=self
        )
        if response:
            self.load_all_dictionaries()
            self.update_statistics()
            self.clear_search()
            
    def count_entries(self, content):
        """Count non-empty, non-comment lines in dictionary content"""
        lines = content.strip().split('\n')
        count = 0
        for line in lines:
            line = line.strip()
            if line and not line.startswith('#'):
                count += 1
        return count
        
    def update_statistics(self):
        """Update statistics display"""
        stats = {}
        total = 0
        
        for key, dict_info in self.dictionaries.items():
            text_widget = self.text_widgets[key]
            content = text_widget.get('1.0', tk.END)
            count = self.count_entries(content)
            stats[key] = count
            total += count
        
        stats_text = (
            f"Словник помилок: {stats['rules']} правил  |  "
            f"Чорний список: {stats['blacklist']} термінів  |  "
            f"Всього: {total} записів"
        )
        self.lbl_stats.config(text=stats_text)
        
    def on_search(self, event=None):
        """Handle search input"""
        search_term = self.entry_search.get().strip().lower()
        
        if not search_term:
            self.clear_highlights()
            return
        
        # Clear previous highlights
        self.clear_highlights()
        
        # Search in current tab using explicit mapping
        current_tab_index = self.notebook.index(self.notebook.select())
        current_key = self.tab_to_key[current_tab_index]
        text_widget = self.text_widgets[current_key]
        
        # Highlight matches
        self.highlight_matches(text_widget, search_term)
        
    def highlight_matches(self, text_widget, search_term):
        """Highlight search matches in text widget"""
        # Configure tag for highlighting
        text_widget.tag_configure('search_highlight', background='yellow', foreground='black')
        
        # Remove previous highlights
        text_widget.tag_remove('search_highlight', '1.0', tk.END)
        
        if not search_term:
            return
        
        # Search and highlight
        start_pos = '1.0'
        first_match = None
        
        while True:
            start_pos = text_widget.search(search_term, start_pos, tk.END, nocase=True)
            if not start_pos:
                break
            
            # Store the first match for scrolling
            if first_match is None:
                first_match = start_pos
            
            end_pos = f"{start_pos}+{len(search_term)}c"
            text_widget.tag_add('search_highlight', start_pos, end_pos)
            start_pos = end_pos
        
        # Scroll to first match if found
        if first_match:
            text_widget.see(first_match)
                
    def clear_highlights(self):
        """Clear all search highlights"""
        for text_widget in self.text_widgets.values():
            text_widget.tag_remove('search_highlight', '1.0', tk.END)
            
    def clear_search(self):
        """Clear search entry and highlights"""
        self.entry_search.delete(0, tk.END)
        self.clear_highlights()


class RowEditDialog(Toplevel):
    """Dialog for adding/editing dictionary rows."""
    
    def __init__(self, parent, columns, values=None, title="Редагувати запис", dictionary_key=None):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title(title)
        self.geometry("700x400")
        self.result = None
        self.columns = columns
        self.dictionary_key = dictionary_key  # Для визначення типу словника
        
        # Create main frame
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill='both', expand=True)
        
        # Create entry fields for each column
        self.entries = {}
        for idx, col in enumerate(columns):
            label = ttk.Label(main_frame, text=f"{col}:", font=('Segoe UI', 10, 'bold'))
            label.grid(row=idx, column=0, sticky='w', pady=5, padx=5)
            
            # Радіокнопки для поля "дія" в чорному списку
            if col == 'дія' and dictionary_key == 'blacklist':
                var = tk.StringVar(value="Видалити")
                
                # Створюємо фрейм для радіокнопок
                radio_frame = ttk.Frame(main_frame)
                radio_frame.grid(row=idx, column=1, sticky='w', pady=5, padx=5)
                
                radio1 = ttk.Radiobutton(radio_frame, text="Видаляти", variable=var, value="Видалити")
                radio1.pack(side='left', padx=(0, 15))
                
                radio2 = ttk.Radiobutton(radio_frame, text="Повідомляти", variable=var, value="Повідомляти")
                radio2.pack(side='left')
                
                self.entries[col] = var
                
                if values and col in values:
                    # Інтерпретуємо значення: "Видалити", "Видаляти", "✓" або будь-яке непорожнє → "Видалити"
                    # "Повідомляти", "Підсвітити" → "Повідомляти"
                    # Порожнє → "Повідомляти"
                    val = values[col].strip()
                    if val in ["Видалити", "Видаляти", "✓"] or (val and val not in ["Повідомляти", "Підсвітити"]):
                        var.set("Видалити")
                    else:
                        var.set("Повідомляти")
            # Use Text widget for multi-line support on larger fields
            elif col in ['коментар', 'виключення', 'українські_корені']:
                text_widget = tk.Text(main_frame, height=3, wrap='word', font=('Consolas', 10))
                text_widget.grid(row=idx, column=1, sticky='ew', pady=5, padx=5)
                self.entries[col] = text_widget
                
                if values and col in values:
                    text_widget.insert('1.0', values[col])
            else:
                entry = ttk.Entry(main_frame, font=('Consolas', 10))
                entry.grid(row=idx, column=1, sticky='ew', pady=5, padx=5)
                self.entries[col] = entry
                
                if values and col in values:
                    entry.insert(0, values[col])
        
        main_frame.columnconfigure(1, weight=1)
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=len(columns), column=0, columnspan=2, pady=10, sticky='e')
        
        ttk.Button(btn_frame, text="Зберегти", command=self.save).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="Скасувати", command=self.destroy).pack(side='right')
        
        # Додаємо підтримку кириличних скорочень
        self.bind_cyrillic_shortcuts()
    
    def bind_cyrillic_shortcuts(self):
        """Прив'язка кириличних клавіатурних скорочень"""
        self.bind('<Control-Key-с>', lambda e: self.handle_copy(e))
        self.bind('<Control-Key-С>', lambda e: self.handle_copy(e))
        self.bind('<Control-Key-м>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-М>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-в>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-В>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-ч>', lambda e: self.handle_cut(e))
        self.bind('<Control-Key-Ч>', lambda e: self.handle_cut(e))
    
    def handle_copy(self, event):
        """Обробка копіювання"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'selection_get'):
                widget.event_generate('<<Copy>>')
        except:
            pass
        return 'break'
    
    def handle_paste(self, event):
        """Обробка вставки"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'insert'):
                widget.event_generate('<<Paste>>')
        except:
            pass
        return 'break'
    
    def handle_cut(self, event):
        """Обробка вирізання"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'selection_get'):
                widget.event_generate('<<Cut>>')
        except:
            pass
        return 'break'
    
    def save(self):
        """Save the edited values."""
        self.result = {}
        for col, widget in self.entries.items():
            if isinstance(widget, tk.StringVar):
                # Для радіокнопок зберігаємо значення напряму
                self.result[col] = widget.get()
            elif isinstance(widget, tk.BooleanVar):
                # Для чекбоксу зберігаємо "✓" якщо увімкнено, інакше порожній рядок
                self.result[col] = "✓" if widget.get() else ""
            elif isinstance(widget, tk.Text):
                self.result[col] = widget.get('1.0', tk.END).strip()
            else:
                self.result[col] = widget.get().strip()
        self.destroy()


class TableDictionaryManager(Toplevel):
    """Tabular dictionary manager with CSV support."""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title("📚 Словники (Табличний формат)")
        self.geometry("1100x650")
        self.parent = parent
        
        # Dictionary configurations
        self.dictionaries = {
            'rules': {
                'file': RULES_FILE_CSV,
                'columns': DEFAULT_RULES_CSV_HEADER,
                'title': '🔄 Помилки перекладу',
                'data': []
            },
            'blacklist': {
                'file': BLACKLIST_FILE_CSV,
                'columns': DEFAULT_BLACKLIST_CSV_HEADER,
                'title': '⛔ Чорний список',
                'data': []
            }
        }
        
        self.create_widgets()
        self.load_all_dictionaries()
        self.update_statistics()
        self.bind_cyrillic_shortcuts()
    
    def bind_cyrillic_shortcuts(self):
        """Прив'язка кириличних клавіатурних скорочень"""
        self.bind('<Control-Key-с>', lambda e: self.handle_copy(e))
        self.bind('<Control-Key-С>', lambda e: self.handle_copy(e))
        self.bind('<Control-Key-м>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-М>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-в>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-В>', lambda e: self.handle_paste(e))
        self.bind('<Control-Key-ч>', lambda e: self.handle_cut(e))
        self.bind('<Control-Key-Ч>', lambda e: self.handle_cut(e))
    
    def handle_copy(self, event):
        """Обробка копіювання"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'selection_get'):
                widget.event_generate('<<Copy>>')
        except:
            pass
        return 'break'
    
    def handle_paste(self, event):
        """Обробка вставки"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'insert'):
                widget.event_generate('<<Paste>>')
        except:
            pass
        return 'break'
    
    def handle_cut(self, event):
        """Обробка вирізання"""
        try:
            widget = self.focus_get()
            if hasattr(widget, 'selection_get'):
                widget.event_generate('<<Cut>>')
        except:
            pass
        return 'break'
        
    def create_widgets(self):
        """Create the table-based UI."""
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill='both', expand=True)
        
        # Statistics frame at the top
        stats_frame = ttk.LabelFrame(main_frame, text="📊 Статистика", padding=5)
        stats_frame.pack(fill='x', pady=(0, 10))
        
        self.lbl_stats = ttk.Label(stats_frame, text="Завантаження...", font=('Segoe UI', 9))
        self.lbl_stats.pack(anchor='w', padx=5)
        
        # Search frame
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(search_frame, text="🔍 Пошук:", font=('Segoe UI', 9, 'bold')).pack(side='left', padx=5)
        self.entry_search = ttk.Entry(search_frame, width=50)
        self.entry_search.pack(side='left', padx=5, fill='x', expand=True)
        self.entry_search.bind('<KeyRelease>', self.on_search)
        
        self.btn_clear_search = ttk.Button(search_frame, text="Очистити", command=self.clear_search)
        self.btn_clear_search.pack(side='left', padx=5)
        
        # Notebook (tabs)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(0, 10))
        self.notebook.bind('<<NotebookTabChanged>>', self.on_tab_changed)
        
        # Create tabs for each dictionary
        self.trees = {}
        self.tab_to_key = {}
        
        for idx, (key, dict_info) in enumerate(self.dictionaries.items()):
            tab_frame = ttk.Frame(self.notebook, padding=5)
            self.notebook.add(tab_frame, text=dict_info['title'])
            self.tab_to_key[idx] = key
            
            # Create Treeview
            tree_frame = ttk.Frame(tab_frame)
            tree_frame.pack(fill='both', expand=True)
            
            columns = dict_info['columns']
            tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode='browse')
            
            # Configure column headings
            for col in columns:
                tree.heading(col, text=col, anchor='w')
                # Set column widths based on content type
                if col in ['російський_корінь', 'українські_корені', 'фраза', 'термін']:
                    tree.column(col, width=150, anchor='w')
                elif col in ['виключення', 'коментар']:
                    tree.column(col, width=250, anchor='w')
                else:
                    tree.column(col, width=120, anchor='w')
            
            # Scrollbars
            vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
            hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            tree.grid(row=0, column=0, sticky='nsew')
            vsb.grid(row=0, column=1, sticky='ns')
            hsb.grid(row=1, column=0, sticky='ew')
            
            tree_frame.columnconfigure(0, weight=1)
            tree_frame.rowconfigure(0, weight=1)
            
            # Double-click to edit
            tree.bind('<Double-1>', lambda e, k=key: self.edit_row(k))
            
            self.trees[key] = tree
            
            # Buttons for this tab
            btn_frame = ttk.Frame(tab_frame)
            btn_frame.pack(fill='x', pady=(5, 0))
            
            ttk.Button(btn_frame, text="➕ Додати", command=lambda k=key: self.add_row(k)).pack(side='left', padx=2)
            ttk.Button(btn_frame, text="✏️ Редагувати", command=lambda k=key: self.edit_row(k)).pack(side='left', padx=2)
            ttk.Button(btn_frame, text="🗑️ Видалити", command=lambda k=key: self.delete_row(k)).pack(side='left', padx=2)
            ttk.Button(btn_frame, text="📤 Експорт в Excel", command=lambda k=key: self.export_to_excel(k)).pack(side='right', padx=2)
        
        # Bottom button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x')
        
        self.btn_save_all = ttk.Button(btn_frame, text="💾 Зберегти всі", command=self.save_all)
        self.btn_save_all.pack(side='left', padx=5)
        
        self.btn_reload = ttk.Button(btn_frame, text="🔄 Перезавантажити", command=self.reload_all)
        self.btn_reload.pack(side='left', padx=5)
        
        self.btn_close = ttk.Button(btn_frame, text="Закрити", command=self.destroy)
        self.btn_close.pack(side='right', padx=5)
    
    def load_all_dictionaries(self):
        """Load content from all CSV dictionary files."""
        for key, dict_info in self.dictionaries.items():
            file_path = dict_info['file']
            
            # Ensure file exists
            if not os.path.exists(file_path):
                create_default_csv_files()
            
            # Load CSV
            try:
                with open(file_path, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.DictReader(f)
                    dict_info['data'] = list(reader)
            except Exception as e:
                messagebox.showerror("Помилка", f"Не вдалося завантажити {file_path}:\n{e}", parent=self)
                dict_info['data'] = []
            
            # Populate tree
            self.populate_tree(key)
    
    def populate_tree(self, key):
        """Populate a tree view with data."""
        tree = self.trees[key]
        dict_info = self.dictionaries[key]
        
        # Clear existing items
        for item in tree.get_children():
            tree.delete(item)
        
        # Add data
        for row in dict_info['data']:
            values = [row.get(col, '') for col in dict_info['columns']]
            tree.insert('', 'end', values=values)
    
    def add_row(self, key):
        """Add a new row to the dictionary."""
        dict_info = self.dictionaries[key]
        columns = dict_info['columns']
        
        dialog = RowEditDialog(self, columns, title=f"Додати запис - {dict_info['title']}", dictionary_key=key)
        self.wait_window(dialog)
        
        if dialog.result:
            dict_info['data'].append(dialog.result)
            self.populate_tree(key)
            self.update_statistics()
    
    def edit_row(self, key):
        """Edit selected row."""
        tree = self.trees[key]
        dict_info = self.dictionaries[key]
        
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Увага", "Виберіть рядок для редагування", parent=self)
            return
        
        item = selection[0]
        idx = tree.index(item)
        current_values = dict_info['data'][idx]
        
        dialog = RowEditDialog(self, dict_info['columns'], current_values, 
                               title=f"Редагувати запис - {dict_info['title']}", dictionary_key=key)
        self.wait_window(dialog)
        
        if dialog.result:
            dict_info['data'][idx] = dialog.result
            self.populate_tree(key)
            self.update_statistics()
    
    def delete_row(self, key):
        """Delete selected row."""
        tree = self.trees[key]
        dict_info = self.dictionaries[key]
        
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Увага", "Виберіть рядок для видалення", parent=self)
            return
        
        response = messagebox.askyesno("Підтвердження", "Видалити вибраний запис?", parent=self)
        if response:
            item = selection[0]
            idx = tree.index(item)
            dict_info['data'].pop(idx)
            self.populate_tree(key)
            self.update_statistics()
    
    def export_to_excel(self, key):
        """Export current dictionary to Excel."""
        dict_info = self.dictionaries[key]
        
        file_path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension='.xlsx',
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{key}_export.xlsx"
        )
        
        if file_path:
            try:
                df = pd.DataFrame(dict_info['data'])
                df.to_excel(file_path, index=False, engine='openpyxl')
                messagebox.showinfo("Успіх", f"Експортовано в {file_path}", parent=self)
            except Exception as e:
                messagebox.showerror("Помилка", f"Не вдалося експортувати:\n{e}", parent=self)
    
    def save_all(self):
        """Save all dictionaries to CSV files."""
        try:
            for key, dict_info in self.dictionaries.items():
                file_path = dict_info['file']
                
                with open(file_path, 'w', encoding='utf-8', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=dict_info['columns'])
                    writer.writeheader()
                    writer.writerows(dict_info['data'])
            
            self.update_statistics()
            messagebox.showinfo("Успіх", "Всі словники успішно збережено!", parent=self)
            
            if hasattr(self.parent, 'log_message'):
                self.parent.log_message("Словники оновлено через Табличний менеджер")
                
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося зберегти словники:\n{e}", parent=self)
    
    def reload_all(self):
        """Reload all dictionaries from CSV files."""
        response = messagebox.askyesno(
            "Підтвердження",
            "Перезавантажити всі словники з файлів? Незбережені зміни будуть втрачені.",
            parent=self
        )
        if response:
            self.load_all_dictionaries()
            self.update_statistics()
            self.clear_search()
    
    def update_statistics(self):
        """Update statistics display."""
        stats = {}
        total = 0
        
        for key, dict_info in self.dictionaries.items():
            count = len(dict_info['data'])
            stats[key] = count
            total += count
        
        stats_text = (
            f"Словник помилок: {stats['rules']} правил  |  "
            f"Чорний список: {stats['blacklist']} термінів  |  "
            f"Всього: {total} записів"
        )
        self.lbl_stats.config(text=stats_text)
    
    def on_search(self, event=None):
        """Handle search input."""
        search_term = self.entry_search.get().strip().lower()
        
        if not search_term:
            # Show all items
            current_tab_index = self.notebook.index(self.notebook.select())
            current_key = self.tab_to_key[current_tab_index]
            self.populate_tree(current_key)
            return
        
        # Filter current tab
        current_tab_index = self.notebook.index(self.notebook.select())
        current_key = self.tab_to_key[current_tab_index]
        tree = self.trees[current_key]
        dict_info = self.dictionaries[current_key]
        
        # Clear tree
        for item in tree.get_children():
            tree.delete(item)
        
        # Add matching rows
        for row in dict_info['data']:
            # Search in all columns
            if any(search_term in str(value).lower() for value in row.values()):
                values = [row.get(col, '') for col in dict_info['columns']]
                tree.insert('', 'end', values=values)
    
    def on_tab_changed(self, event=None):
        """Handle tab change - refresh search."""
        self.on_search()
    
    def clear_search(self):
        """Clear search and show all items."""
        self.entry_search.delete(0, tk.END)
        for key in self.dictionaries:
            self.populate_tree(key)


class SpellCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"TextChecker {APP_VERSION}")
        self.root.geometry("980x700")
        self.root.minsize(920, 640)
        self.style = ttk.Style(self.root)
        self.style.theme_use('clam')
        self.configure_styles()

        self.file_paths = []
        self.last_result_paths = []
        # Режим обробки: користувач може вибрати "Тільки очищення HTML"
        self.skip_processed_var = tk.BooleanVar(value=False)  # Не використовується
        self.html_only_var = tk.BooleanVar(value=False)  # За замовчуванням вимкнено
        
        # Perform migration from TXT to CSV if needed
        migrate_txt_to_csv()
        
        # Create default CSV files if they don't exist
        create_default_csv_files()
        
        self.create_widgets()

        # Legacy TXT file creation (for backward compatibility during transition)
        if not os.path.exists(RULES_FILE):
            with open(RULES_FILE, 'w', encoding='utf-8') as f:
                f.write(DEFAULT_RULES)
        if not os.path.exists(IGNORE_FILE):
            with open(IGNORE_FILE, 'w', encoding='utf-8') as f:
                f.write(DEFAULT_IGNORES)
        if not os.path.exists(BLACKLIST_FILE):
            with open(BLACKLIST_FILE, 'w', encoding='utf-8') as f:
                f.write(DEFAULT_BLACKLIST)

    def configure_styles(self):
        self.style.configure('Title.TLabel', font=('Segoe UI', 12, 'bold'))
        self.style.configure('TLabelframe', padding=8)
        self.style.configure('TLabelframe.Label', font=('Segoe UI', 10, 'bold'))
        self.style.configure('TButton', padding=(10, 6))
        self.style.configure('TCheckbutton', padding=2)
        
        # Configure checkbutton to use checkmark instead of cross
        # Create custom images for checked/unchecked states with checkmarks
        try:
            # Checkmark coordinates
            CHECKMARK_STEM_X1, CHECKMARK_STEM_Y1 = 5, 8
            CHECKMARK_STEM_X2, CHECKMARK_STEM_Y2 = 7, 10
            CHECKMARK_MID_X1, CHECKMARK_MID_Y1 = 7, 10
            CHECKMARK_MID_X2, CHECKMARK_MID_Y2 = 9, 12
            CHECKMARK_TOP_X1, CHECKMARK_TOP_Y1 = 9, 6
            CHECKMARK_TOP_X2, CHECKMARK_TOP_Y2 = 11, 8
            
            # Small checkmark icon for checked state (✓)
            checked_img = tk.PhotoImage(width=16, height=16)
            checked_img.put(("#4CAF50",) * 16, to=(0, 0, 16, 16))  # Green background
            checked_img.put(("#FFFFFF",) * 2, to=(CHECKMARK_STEM_X1, CHECKMARK_STEM_Y1, CHECKMARK_STEM_X2, CHECKMARK_STEM_Y2))
            checked_img.put(("#FFFFFF",) * 2, to=(CHECKMARK_MID_X1, CHECKMARK_MID_Y1, CHECKMARK_MID_X2, CHECKMARK_MID_Y2))
            checked_img.put(("#FFFFFF",) * 2, to=(CHECKMARK_TOP_X1, CHECKMARK_TOP_Y1, CHECKMARK_TOP_X2, CHECKMARK_TOP_Y2))
            
            # Empty box for unchecked state
            unchecked_img = tk.PhotoImage(width=16, height=16)
            unchecked_img.put(("#E0E0E0",) * 16, to=(0, 0, 16, 16))  # Light gray background
            unchecked_img.put(("#FFFFFF",) * 14, to=(1, 1, 15, 15))  # White interior
            
            # Keep references to prevent garbage collection
            self.checked_img = checked_img
            self.unchecked_img = unchecked_img
            
            # Create custom element with images
            self.style.element_create('custom.indicator', 'image', unchecked_img,
                                     ('selected', checked_img), width=16, border=0, sticky='w')
            
            # Update layout to use custom indicator
            self.style.layout('TCheckbutton',
                [('Checkbutton.padding',
                  {'children': [('custom.indicator', {'side': 'left', 'sticky': ''}),
                                ('Checkbutton.label', {'side': 'left', 'sticky': 'nswe'})],
                   'sticky': 'nswe'})])
        except (tk.TclError, AttributeError) as e:
            # If custom styling fails (e.g., due to theme incompatibility), fall back to default
            # Log error for debugging but don't crash the application
            print(f"Warning: Could not apply custom checkbox styling: {e}")

    def create_widgets(self):
        ttk.Label(self.root, text=f"TextChecker {APP_VERSION}", style='Title.TLabel').pack(
            anchor='w', padx=14, pady=(12, 4)
        )

        file_frame = ttk.LabelFrame(self.root, text="1. Вибір файлів та колонок")
        file_frame.pack(fill='x', padx=10, pady=5)

        file_select_row = ttk.Frame(file_frame)
        file_select_row.pack(fill='x', pady=5)

        self.btn_select_file = ttk.Button(file_select_row, text="📂 Обрати Excel-файли", command=self.select_files)
        self.btn_select_file.pack(side='left', padx=5)

        self.lbl_file_status = ttk.Label(file_select_row, text="Файли не обрано", foreground="gray")
        self.lbl_file_status.pack(side='left', padx=8, fill='x', expand=True)

        columns_row = ttk.Frame(file_frame)
        columns_row.pack(fill='x', pady=(8, 4))

        ttk.Label(columns_row, text="Стовпчик RU:", font=('Arial', 9, 'bold')).pack(side='left', padx=5)
        self.combo_ru = ttk.Combobox(columns_row, state="disabled", width=25)
        self.combo_ru.pack(side='left', padx=5)

        ttk.Label(columns_row, text="Стовпчик UA:", font=('Arial', 9, 'bold')).pack(side='left', padx=15)
        self.combo_ua = ttk.Combobox(columns_row, state="disabled", width=25)
        self.combo_ua.pack(side='left', padx=5)

        # Секція режимів обробки
        options_frame = ttk.LabelFrame(self.root, text="2. Режим обробки")
        options_frame.pack(fill='x', padx=10, pady=5)

        self.chk_html_only = ttk.Checkbutton(
            options_frame,
            text="Тільки очищення HTML (без перевірки правил перекладу)",
            variable=self.html_only_var
        )
        self.chk_html_only.pack(side='left', padx=5)

        control_frame = ttk.LabelFrame(self.root, text="3. Керування та словники")
        control_frame.pack(fill='x', padx=10, pady=5)

        self.btn_start_analysis = ttk.Button(control_frame, text="▶ Почати перевірку", command=self.start_analysis, state='disabled')
        self.btn_start_analysis.pack(side='left', padx=5, fill='x', expand=True)

        self.btn_open_result = ttk.Button(control_frame, text="📂 Відкрити результат", command=self.open_results, state='disabled')
        self.btn_open_result.pack(side='left', padx=5)

        self.btn_dictionary_manager = ttk.Button(control_frame, text="📚 Словники", command=self.open_dictionary_manager)
        self.btn_dictionary_manager.pack(side='left', padx=5)

        progress_frame = ttk.Frame(self.root, padding="5")
        progress_frame.pack(fill='x', padx=10, pady=5)

        self.progress_bar = ttk.Progressbar(progress_frame, orient='horizontal', mode='determinate', length=100)
        self.progress_bar.pack(fill='x', expand=True)

        self.lbl_progress_status = ttk.Label(progress_frame, text="Очікування файлу...", anchor='center')
        self.lbl_progress_status.pack(fill='x', pady=2)

        self.txt_log = tk.Text(self.root, height=10, state='disabled', wrap='word', bg="#f0f0f0", font=('Consolas', 9))
        self.txt_log.pack(fill='both', expand=True, padx=10, pady=10)

        log_scroll = ttk.Scrollbar(self.txt_log, orient='vertical', command=self.txt_log.yview)
        self.txt_log['yscrollcommand'] = log_scroll.set
        log_scroll.pack(side='right', fill='y')
        
        # Додаємо підтримку кириличних клавіш для копіювання/вставки
        self.bind_cyrillic_shortcuts()
    
    def bind_cyrillic_shortcuts(self):
        """Прив'язка кириличних клавіатурних скорочень для копіювання/вставки"""
        # Ctrl+С (кирилична С) = копіювати
        self.root.bind('<Control-Key-с>', lambda e: self.handle_copy(e))
        self.root.bind('<Control-Key-С>', lambda e: self.handle_copy(e))
        
        # Ctrl+М (кирилична М) = вставити (російська В на тій же клавіші що і V)
        self.root.bind('<Control-Key-м>', lambda e: self.handle_paste(e))
        self.root.bind('<Control-Key-М>', lambda e: self.handle_paste(e))
        
        # Також Ctrl+В (кирилична В) = вставити
        self.root.bind('<Control-Key-в>', lambda e: self.handle_paste(e))
        self.root.bind('<Control-Key-В>', lambda e: self.handle_paste(e))
        
        # Ctrl+Ч (кирилична Ч на клавіші X) = вирізати
        self.root.bind('<Control-Key-ч>', lambda e: self.handle_cut(e))
        self.root.bind('<Control-Key-Ч>', lambda e: self.handle_cut(e))
        
        # Ctrl+Ш (кирилична Ш на клавіші I для Insert) - вставити
        self.root.bind('<Control-Key-ш>', lambda e: self.handle_paste(e))
        self.root.bind('<Control-Key-Ш>', lambda e: self.handle_paste(e))
    
    def handle_copy(self, event):
        """Обробка копіювання з кириличної розкладки"""
        try:
            widget = self.root.focus_get()
            if hasattr(widget, 'selection_get'):
                widget.event_generate('<<Copy>>')
        except:
            pass
        return 'break'
    
    def handle_paste(self, event):
        """Обробка вставки з кириличної розкладки"""
        try:
            widget = self.root.focus_get()
            if hasattr(widget, 'insert'):
                widget.event_generate('<<Paste>>')
        except:
            pass
        return 'break'
    
    def handle_cut(self, event):
        """Обробка вирізання з кириличної розкладки"""
        try:
            widget = self.root.focus_get()
            if hasattr(widget, 'selection_get'):
                widget.event_generate('<<Cut>>')
        except:
            pass
        return 'break'

    def log_message(self, message):
        self.txt_log.config(state='normal')
        timestamp = time.strftime("%H:%M:%S")
        self.txt_log.insert(tk.END, f"[{timestamp}] {message}\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state='disabled')

    def open_editor(self, title, file_path, default_content):
        editor = TextEditor(self.root, title, file_path, default_content)
        self.root.wait_window(editor)
        self.log_message(f"Файл '{title}' оновлено.")
    
    def open_dictionary_manager(self):
        """Open the table-based dictionary manager window"""
        manager = TableDictionaryManager(self.root)
        self.root.wait_window(manager)

    def detect_default_columns(self, columns):
        normalized = {col: str(col).strip().lower() for col in columns}
        ru_priority = ['описание;1', 'описание', 'description;1', 'description ru', 'ru', 'рус']
        ua_priority = ['описание (ua);1', 'опис (ua);1', 'описание ua;1', 'описание (ua)', 'description (ua);1', 'description ua', 'ua', 'uk', 'укр']

        def find_by_priority(priority_list):
            for target in priority_list:
                for original, lowered in normalized.items():
                    if lowered == target:
                        return original
            return ''

        ru_col = find_by_priority(ru_priority)
        ua_col = find_by_priority(ua_priority)

        if not ru_col:
            ru_col = next((c for c in columns if any(token in normalized[c] for token in ['описание', 'description']) and '(ua)' not in normalized[c] and 'ua' not in normalized[c] and 'uk' not in normalized[c]), '')
        if not ua_col:
            ua_col = next((c for c in columns if any(token in normalized[c] for token in ['ua', 'uk', 'укр', '(ua)'])), '')

        return ru_col, ua_col

    def update_file_selection(self, paths):
        if not paths:
            return
        ordered_paths = list(dict.fromkeys(paths))
        self.file_paths = ordered_paths
        try:
            df_preview = pd.read_excel(ordered_paths[0], nrows=0)
            columns = df_preview.columns.tolist()
            self.combo_ru.config(values=columns, state="readonly")
            self.combo_ua.config(values=columns, state="readonly")
            ru_col, ua_col = self.detect_default_columns(columns)
            if ru_col:
                self.combo_ru.set(ru_col)
            if ua_col:
                self.combo_ua.set(ua_col)
            if len(ordered_paths) == 1:
                self.lbl_file_status.config(text=os.path.basename(ordered_paths[0]), foreground="green")
            else:
                self.lbl_file_status.config(
                    text=f"Обрано файлів: {len(ordered_paths)} (перший: {os.path.basename(ordered_paths[0])})",
                    foreground="green"
                )
            self.log_message(f"Завантажено {len(ordered_paths)} файл(ів). Можна починати перевірку.")
            self.btn_start_analysis.config(state='normal')
        except Exception as e:
            messagebox.showerror("Помилка файлу", f"Не вдалося зчитати файл:\n{e}")

    def select_files(self):
        paths = filedialog.askopenfilenames(
            title="Оберіть один або кілька файлів Excel",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        self.update_file_selection(list(paths))

    def set_ui_state(self, is_running):
        state = 'disabled' if is_running else 'normal'
        self.btn_select_file.config(state=state)
        self.btn_dictionary_manager.config(state=state)
        start_state = state if self.file_paths else 'disabled'
        self.btn_start_analysis.config(state=start_state)
        self.chk_skip_processed.config(state=state)
        self.chk_html_only.config(state=state)
        self.combo_ru.config(state="disabled" if is_running else "readonly")
        self.combo_ua.config(state="disabled" if is_running else "readonly")
        if is_running:
            self.btn_open_result.config(state='disabled')

    def parse_rules(self):
        """Parse translation rules from CSV or TXT format."""
        rules = {}
        
        # Try CSV first (new format)
        if os.path.exists(RULES_FILE_CSV):
            try:
                with open(RULES_FILE_CSV, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        ru_stem = row.get('російський_корінь', '').strip()
                        ua_stems_str = row.get('українські_корені', '').strip()
                        
                        if ru_stem and ua_stems_str:
                            ua_stems = [ua.strip() for ua in ua_stems_str.split(',')]
                            rules[ru_stem] = ua_stems
                return rules
            except Exception as e:
                print(f"Помилка читання CSV: {e}, перемикаємось на TXT")
        
        # Fallback to TXT format (legacy)
        if os.path.exists(RULES_FILE):
            with open(RULES_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.split('#')[0].strip()
                    if '=' in line:
                        ru, ua_list = line.split('=', 1)
                        ua_stems = [ua.strip() for ua in ua_list.split(',')]
                        rules[ru.strip()] = ua_stems
        
        return rules

    def parse_ignores(self):
        """Parse ignore rules from CSV or TXT format."""
        ignores = []
        
        # Try CSV first (new format)
        if os.path.exists(IGNORE_FILE_CSV):
            try:
                with open(IGNORE_FILE_CSV, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        phrase = row.get('фраза', '').strip()
                        if phrase:
                            ignores.append(phrase.lower())
                return ignores
            except Exception as e:
                print(f"Помилка читання CSV: {e}, перемикаємось на TXT")
        
        # Fallback to TXT format (legacy)
        if os.path.exists(IGNORE_FILE):
            with open(IGNORE_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.split('#')[0].strip()
                    if line:
                        ignores.append(line.lower())
        
        return ignores

    def parse_blacklist(self):
        """
        Parse blacklist from CSV or TXT format.
        Returns: list of dicts with 'term', 'exceptions', 'action' keys
        """
        blacklist = []
        
        # Try CSV first (new format)
        if os.path.exists(BLACKLIST_FILE_CSV):
            try:
                with open(BLACKLIST_FILE_CSV, 'r', encoding='utf-8', newline='') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        term = row.get('термін', '').strip()
                        if term:
                            exceptions_str = row.get('виключення', '').strip()
                            exceptions = [e.strip().lower() for e in exceptions_str.split(',') if e.strip()]
                            action = row.get('дія', 'Видалити').strip()
                            
                            blacklist.append({
                                'term': term.lower(),
                                'exceptions': exceptions,
                                'action': action
                            })
                return blacklist
            except Exception as e:
                print(f"Помилка читання CSV: {e}, перемикаємось на TXT")
        
        # Fallback to TXT format (legacy) - without exceptions and with default action
        if os.path.exists(BLACKLIST_FILE):
            with open(BLACKLIST_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.split('#')[0].strip()
                    if line:
                        blacklist.append({
                            'term': line.lower(),
                            'exceptions': [],
                            'action': 'Видалити'
                        })
        
        return blacklist

    def normalize_text(self, text):
        if pd.isna(text):
            return ""
        return self.normalize_common_text_issues(str(text))

    def normalize_common_text_issues(self, text):
        text = str(text).replace('\u00A0', ' ').replace('\u200B', '')
        text = text.replace('“', '"').replace('”', '"').replace('„', '"')
        text = text.replace('’', "'").replace('`', "'")
        # Convert decimal commas to dots only in likely measurement/value contexts.
        text = re.sub(
            rf'(?<!\d)(\d{{1,{DECIMAL_MAX_INT_DIGITS}}}),(\d{{1,{DECIMAL_MAX_FRACTION_DIGITS}}})(?=\s*(?:{DECIMAL_CONTEXT_UNITS}|$|[)\].,;:!?]))',
            r'\1.\2',
            text,
            flags=re.IGNORECASE
        )
        text = re.sub(r'[ \t]+', ' ', text)
        return text.strip()

    def strip_technical_html(self, text):
        text = self.normalize_text(text)
        text = unescape(text)
        text = re.sub(r'\[/?html\]', '', text, flags=re.IGNORECASE)
        for pattern in TECHNICAL_HTML_PATTERNS:
            text = re.sub(pattern, ' ', text)
        text = re.sub(r'(?is)<!--.*?-->', ' ', text)
        return text

    def clean_html_text(self, text):
        text = self.strip_technical_html(text)
        text = re.sub(r'(?is)<style\b[^>]*>.*?</style>', ' ', text)
        text = re.sub(r'(?is)<script\b[^>]*>.*?</script>', ' ', text)
        text = re.sub(r'(?i)(</li>|<br\s*/?>|</p>|</div>)', '. ', text)
        text = re.sub(r'<[^>]+>', ' ', text)
        text = re.sub(r'http[s]?://\S+', ' ', text)
        text = re.sub(r'\b[a-z0-9_-]+\s*\{[^{}]*\}', ' ', text, flags=re.IGNORECASE)
        text = re.sub(r'\s+', ' ', text).strip(' .')
        return text

    def is_in_technical_context(self, text, match_pos):
        """Check if a position in text is within technical HTML/URL context."""
        # Check if match is inside HTML attributes (src, href, cssurl, poster, data-*, background-image, url())
        before_match = text[:match_pos].lower()
        after_match = text[match_pos:].lower()
        
        # Check for attribute contexts
        # These patterns check if we're right after an attribute assignment or url( function
        attr_patterns = [
            r'src\s*=\s*["\']?[^"\'>\s]*$',
            r'href\s*=\s*["\']?[^"\'>\s]*$',
            r'cssurl\s*=\s*["\']?[^"\'>\s]*$',
            r'poster\s*=\s*["\']?[^"\'>\s]*$',
            r'data-[a-z-]+\s*=\s*["\']?[^"\'>\s]*$',
            r'background-image\s*:\s*url\(["\']?[^"\')\s]*$',
            r'url\(["\']?[^"\')\s]*$',
            r'background\s*:\s*url\(["\']?[^"\')\s]*$',
            r'content\s*:\s*url\(["\']?[^"\')\s]*$',
        ]
        
        for pattern in attr_patterns:
            if re.search(pattern, before_match):
                # We found an attribute pattern before the match position
                # This indicates we're inside a URL or attribute value
                
                # Check if the match looks like a URL continuation (allowing uppercase, query params, fragments, etc.)
                # This is the primary check for URL-like content
                if re.match(r'^https?://[a-zA-Z0-9._/\-?#&=+%@]+', after_match):
                    # Looks like a valid URL - we're in technical context
                    return True
                
                # Secondary check: if we're continuing inside an attribute value
                # (no immediate whitespace or new tag start) and not at end of string
                # This handles edge cases where URL detection above might be too strict
                if after_match and not re.match(r'^[\s<]', after_match):
                    # No immediate whitespace or tag start, likely still in the attribute
                    return True
                
                # End of string after finding an attribute pattern
                if not after_match:
                    return True
        
        # Check if we're inside an HTML tag
        last_open = before_match.rfind('<')
        last_close = before_match.rfind('>')
        if last_open > last_close:
            # We're inside a tag
            next_close = after_match.find('>')
            if next_close != -1:
                return True
        
        # Check for protocol patterns at the start
        if re.search(r'(?:^|["\'\s(])(https?://|ftp://|//)$', before_match):
            return True
        
        return False

    def contains_technical_html(self, text):
        raw_text = self.normalize_text(text)
        if not raw_text:
            return False
        lowered = raw_text.lower()
        if '[html]' in lowered or 'ssd-module-wrap' in lowered:
            return True
        return any(re.search(pattern, raw_text) for pattern in TECHNICAL_HTML_PATTERNS)

    def contains_chinese(self, text):
        return bool(CHINESE_CHAR_PATTERN.search(self.normalize_text(text)))

    def is_meaningful_text(self, text):
        cleaned = self.clean_html_text(text)
        if not cleaned:
            return False
        if self.contains_chinese(cleaned):
            return False
        words = re.findall(r'[A-Za-zА-Яа-яІіЇїЄєҐґЁёЪъЫыЭэ0-9]+', cleaned)
        if len(words) <= 1:
            return False
        if len(cleaned) < 12:
            return False
        if 'ssd-module-wrap' in cleaned.lower() or 'background-image' in cleaned.lower():
            return False
        return True

    def is_suspiciously_short_text(self, text):
        cleaned = self.clean_html_text(text)
        if not cleaned:
            return False
        words = re.findall(r'[A-Za-zА-Яа-яІіЇїЄєҐґЁёЪъЫыЭэ]+', cleaned)
        return len(words) <= MIN_SHORT_DESC_WORDS or len(cleaned) < MIN_SHORT_DESC_LENGTH

    def extract_error_sentence(self, text, error_word_stem):
        text_clean = self.clean_html_text(text)
        sentences = re.split(r'(?<=[.!?\n])\s+', text_clean)
        for s in sentences:
            if re.search(r'(?i)\b' + re.escape(error_word_stem), s):
                return s.strip(' .-;:,\t')
        match = re.search(r'(.{0,50}\b' + re.escape(error_word_stem) + r'.{0,50})', text_clean, re.IGNORECASE)
        if match:
            return "..." + match.group(1).strip() + "..."
        return ""

    def sentence_contains_blacklist(self, sentence, blacklist):
        """
        Check if sentence contains blacklist term.
        Args:
            sentence: text to check
            blacklist: list of dicts with 'term', 'exceptions', 'action' keys
        Returns:
            dict with 'found' (bool), 'term' (str), 'action' (str) or None if no match
        """
        sentence_lower = sentence.lower()
        for entry in blacklist:
            term = entry['term']
            if term in sentence_lower:
                # Check exceptions
                exceptions = entry.get('exceptions', [])
                is_exception = any(exc in sentence_lower for exc in exceptions if exc)
                
                if not is_exception:
                    return {
                        'found': True,
                        'term': term,
                        'action': entry.get('action', 'Видалити')
                    }
        
        return None

    def convert_markdown_bold_to_html(self, text):
        return re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', text)

    def normalize_label(self, label):
        label = re.sub(r'\s+', ' ', label).strip()
        return label[:-1].strip() if label.endswith(':') else label

    def extract_paragraphs(self, source):
        paragraphs = re.findall(r'(?is)<p\b[^>]*>(.*?)</p>', source)
        normalized_paragraphs = []
        for paragraph in paragraphs:
            paragraph = re.sub(r'(?is)<img\b[^>]*>', ' ', paragraph)
            paragraph = self.convert_markdown_bold_to_html(paragraph)
            paragraph = re.sub(r'<br\s*/?>', ' ', paragraph, flags=re.IGNORECASE)
            paragraph = re.sub(r'\s+', ' ', paragraph).strip()
            paragraph_text = self.normalize_common_text_issues(unescape(re.sub(r'<[^>]+>', '', paragraph)).strip())
            if paragraph_text:
                normalized_paragraphs.append((paragraph, paragraph_text))
        if not normalized_paragraphs:
            # Fallback for plain-text inputs without <p> tags.
            plain_text = self.clean_html_text(source)
            if plain_text:
                normalized_paragraphs.append((plain_text, plain_text))
        return normalized_paragraphs

    def classify_paragraph(self, paragraph_html, paragraph_text):
        strong_match = re.match(r'^<strong>(.+?)</strong>\s*(.*)$', paragraph_html, flags=re.IGNORECASE | re.DOTALL)
        if strong_match:
            raw_label = strong_match.group(1).strip()
            value = strong_match.group(2).strip()
            label = self.normalize_label(raw_label)
            if value:
                return {'type': 'labeled_value', 'label': label, 'value': value, 'text': paragraph_text}
            return {'type': 'heading', 'label': label, 'text': paragraph_text}

        # Check for dash-prefixed items first, strip the bullet marker, then re-classify
        # Supporting various bullet markers: -, •, *, —, en-dash (\u2013), em-dash (\u2014)
        # Note: hyphen escaped explicitly for clarity
        bullet_match = re.match(r'^[\u2022\*\u2013\u2014\\\-]\s+(.+)$', paragraph_text)
        if bullet_match:
            # Strip the bullet marker and get the content
            content_after_bullet = bullet_match.group(1).strip()
            
            # Check if the content after the bullet has a colon pattern (label: value)
            colon_match_after_bullet = re.match(r'^([^:]{1,80}):\s*(.+)$', content_after_bullet)
            if colon_match_after_bullet:
                label = self.normalize_label(colon_match_after_bullet.group(1))
                value = colon_match_after_bullet.group(2).strip()
                if label and value:
                    # This is a dash-prefixed labeled item: "- Label: value"
                    return {'type': 'list_item_labeled', 'label': label, 'value': value, 'text': paragraph_text}
            
            # If no colon, it's a simple list item: "- simple text"
            if content_after_bullet:
                return {'type': 'list_item', 'value': content_after_bullet, 'text': paragraph_text}

        # Check for labeled items without dash prefix
        colon_match = re.match(r'^([^:]{1,80}):\s*(.+)$', paragraph_text)
        if colon_match:
            label = self.normalize_label(colon_match.group(1))
            value = colon_match.group(2).strip()
            if label and value:
                return {'type': 'plain_labeled_value', 'label': label, 'value': value, 'text': paragraph_text}

        return {'type': 'text', 'text': paragraph_text}

    def detect_language(self, text):
        cleaned = self.clean_html_text(text).lower()
        if not cleaned:
            return 'unknown'
        if self.contains_chinese(cleaned):
            return 'zh'

        ua_score = sum(cleaned.count(ch) for ch in UKRAINIAN_UNIQUE_CHARS)
        ru_score = sum(cleaned.count(ch) for ch in RUSSIAN_UNIQUE_CHARS)

        words = re.findall(r'[а-яіїєґёъыэ]+', cleaned, flags=re.IGNORECASE)
        for word in words:
            if word in UKRAINIAN_MARKERS:
                ua_score += 2
            if word in RUSSIAN_MARKERS:
                ru_score += 2

        if ua_score == 0 and ru_score == 0:
            return 'unknown'
        if ua_score > ru_score:
            return 'ua'
        if ru_score > ua_score:
            return 'ru'
        return 'unknown'

    def format_description(self, text, blacklist):
        source = self.strip_technical_html(text)
        paragraphs = self.extract_paragraphs(source)
        if not paragraphs:
            return ""

        result = []
        current_list_items = []
        current_list_signatures = set()

        def flush_list():
            nonlocal current_list_items, current_list_signatures
            if current_list_items:
                result.append('<ul>')
                result.extend(current_list_items)
                result.append('</ul>')
                current_list_items = []
                current_list_signatures = set()

        for paragraph_html, paragraph_text in paragraphs:
            blacklist_match = self.sentence_contains_blacklist(paragraph_text, blacklist)
            # Only skip (delete) if action is "Видалити"
            if blacklist_match and blacklist_match.get('action') == 'Видалити':
                continue
            if self.contains_chinese(paragraph_text):
                continue

            item = self.classify_paragraph(paragraph_html, paragraph_text)
            item_type = item['type']

            if item_type == 'heading':
                flush_list()
                result.append(f"<p><strong>{item['label']}:</strong></p>")
                continue

            if item_type in ('labeled_value', 'plain_labeled_value', 'list_item_labeled'):
                label = item['label']
                value = self.normalize_common_text_issues(item['value'])

                if self.contains_chinese(label) or self.contains_chinese(value):
                    continue

                if label.lower() in ('увага', 'примітка', 'внимание', 'примечание'):
                    flush_list()
                    result.append(f"<p><strong>{label}:</strong> {value}</p>")
                    continue

                # Exact-match deduplication (label + value) to avoid dropping similar-but-different items.
                item_signature = (label, value)
                if item_signature not in current_list_signatures:
                    current_list_items.append(f"<li><strong>{label}:</strong> {value}</li>")
                    current_list_signatures.add(item_signature)
                continue

            if item_type == 'list_item':
                value = self.normalize_common_text_issues(item['value'])
                
                if self.contains_chinese(value):
                    continue
                
                # Deduplication for plain list items (by value only)
                # Using None as label placeholder to distinguish from labeled items
                item_signature = (None, value)
                if item_signature not in current_list_signatures:
                    current_list_items.append(f"<li>{value}</li>")
                    current_list_signatures.add(item_signature)
                continue

            flush_list()
            result.append(f"<p>{self.normalize_common_text_issues(item['text'])}</p>")

        flush_list()
        formatted = self.cleanup_formatted_html('\n'.join(result).strip())
        return formatted if self.is_meaningful_text(formatted) else ""

    def cleanup_formatted_html(self, html):
        cleaned = re.sub(r'(?is)<p>\s*(?:&nbsp;|\s)*</p>', '', html)
        cleaned = re.sub(r'(?is)<div[^>]*>\s*(?:&nbsp;|\s)*</div>', '', cleaned)
        cleaned = re.sub(r'\n{2,}', '\n', cleaned)
        return cleaned.strip()

    def get_unique_column_name(self, columns, base_name):
        if base_name not in columns:
            return base_name
        counter = 2
        while f"{base_name}_{counter}" in columns:
            counter += 1
        return f"{base_name}_{counter}"

    def get_existing_or_create_column(self, df, column_name):
        if column_name not in df.columns:
            df[column_name] = ''
        return column_name

    def resolve_columns_for_file(self, columns, preferred_ru, preferred_ua):
        ru_col = preferred_ru if preferred_ru in columns else ''
        ua_col = preferred_ua if preferred_ua in columns else ''
        if not ru_col or not ua_col:
            detected_ru, detected_ua = self.detect_default_columns(columns)
            ru_col = ru_col or detected_ru
            ua_col = ua_col or detected_ua
        if not ru_col or not ua_col:
            raise ValueError("Не вдалося автоматично визначити колонки RU/UA у файлі.")
        return ru_col, ua_col

    def compose_status(self, statuses):
        ordered = []
        for status in statuses:
            if status and status not in ordered:
                ordered.append(status)
        return " | ".join(ordered) if ordered else "OK"

    def build_summary_text(self, summaries):
        total_files = len(summaries)
        totals = {
            'total_rows': 0,
            'processed_rows': 0,
            'rows_with_errors': 0,
            'need_translation': 0,
            'technical_html': 0,
            'blacklist': 0,
            'chinese': 0,
            'short': 0,
            'skipped': 0,
            'ok_rows': 0
        }
        for stats in summaries:
            for key in totals:
                totals[key] += stats.get(key, 0)
        return (
            f"Оброблено файлів: {total_files}\n"
            f"Усього рядків: {totals['total_rows']}\n"
            f"Опрацьовано рядків: {totals['processed_rows']}\n"
            f"Рядків з помилками: {totals['rows_with_errors']}\n"
            f"Потрібен переклад: {totals['need_translation']}\n"
            f"Технічний HTML: {totals['technical_html']}\n"
            f"Заборонений термін: {totals['blacklist']}\n"
            f"Китайський текст: {totals['chinese']}\n"
            f"Підозріло короткий опис: {totals['short']}\n"
            f"Пропущено (вже оброблено): {totals['skipped']}\n"
            f"OK: {totals['ok_rows']}"
        )

    def process_single_file(self, file_path, preferred_ru, preferred_ua, rules, ignores, blacklist, html_only_mode, skip_processed):
        self.log_message(f"Обробка файлу: {os.path.basename(file_path)}")
        df = pd.read_excel(file_path)
        total_rows = len(df)
        self.progress_bar.config(maximum=max(total_rows, 1), value=0)

        col_ru, col_ua = self.resolve_columns_for_file(df.columns, preferred_ru, preferred_ua)
        errors_col_name = self.get_existing_or_create_column(df, 'Помилки')
        status_col_name = self.get_existing_or_create_column(df, 'Статус')
        checked_ru_col_name = self.get_existing_or_create_column(df, f'{col_ru}_checked')
        checked_ua_col_name = self.get_existing_or_create_column(df, f'{col_ua}_checked')

        rows_with_errors = set()
        rows_with_language_mismatch = set()
        rows_need_translation = set()
        rows_with_technical_html = set()
        rows_with_blacklist = set()
        rows_with_chinese = set()
        rows_with_short_desc = set()
        rows_skipped = set()
        rows_ok = set()

        for i, row in df.iterrows():
            text_ru = self.normalize_text(row.get(col_ru, ''))
            text_ua = self.normalize_text(row.get(col_ua, ''))
            row_errors = []
            row_statuses = []

            existing_ru_checked = self.normalize_text(row.get(checked_ru_col_name, ''))
            existing_ua_checked = self.normalize_text(row.get(checked_ua_col_name, ''))
            # Skip when at least one checked field is already populated to avoid accidental overwrite.
            if skip_processed and (existing_ru_checked.strip() or existing_ua_checked.strip()):
                rows_skipped.add(i)
                df.at[i, status_col_name] = "Пропущено (вже оброблено)"
                continue

            ru_detected_lang = self.detect_language(text_ru)
            ua_detected_lang = self.detect_language(text_ua)
            ru_has_technical = self.contains_technical_html(text_ru)
            ua_has_technical = self.contains_technical_html(text_ua)
            ru_has_chinese = self.contains_chinese(text_ru)
            ua_has_chinese = self.contains_chinese(text_ua)

            # Track technical HTML for statistics, but don't add to status
            if ru_has_technical or ua_has_technical:
                rows_with_technical_html.add(i)
            if ru_has_chinese or ua_has_chinese:
                row_statuses.append("Китайський текст")
                rows_with_chinese.add(i)

            if text_ru.strip() and ru_detected_lang in ('ua', 'zh') and not html_only_mode:
                if ru_detected_lang == 'zh':
                    row_errors.append("[МОВА] У стовпчику RU виявлено китайський текст. Потрібен переклад російською.")
                else:
                    row_errors.append("[МОВА] У стовпчику RU виявлено український текст. Потрібен переклад російською.")
                formatted_ru_text = ""
                row_statuses.append("Потрібен переклад")
                rows_need_translation.add(i)
                rows_with_errors.add(i)
                rows_with_language_mismatch.add(i)
            else:
                formatted_ru_text = self.format_description(text_ru, blacklist)
                if text_ru.strip() and not formatted_ru_text:
                    row_errors.append("[КОНТЕНТ] У стовпчику RU не виявлено придатного текстового опису або знайдено технічний HTML/китайський контент.")
                    rows_with_errors.add(i)
                if text_ru.strip() and self.is_suspiciously_short_text(text_ru):
                    row_statuses.append("Підозріло короткий опис")
                    rows_with_short_desc.add(i)

            if text_ua.strip() and ua_detected_lang in ('ru', 'zh') and not html_only_mode:
                if ua_detected_lang == 'zh':
                    row_errors.append("[МОВА] У стовпчику UA виявлено китайський текст. Потрібен переклад українською.")
                else:
                    row_errors.append("[МОВА] У стовпчику UA виявлено російський текст. Потрібен переклад українською.")
                formatted_ua_text = ""
                row_statuses.append("Потрібен переклад")
                rows_need_translation.add(i)
                rows_with_errors.add(i)
                rows_with_language_mismatch.add(i)
            else:
                formatted_ua_text = self.format_description(text_ua, blacklist)
                if text_ua.strip() and not formatted_ua_text:
                    row_errors.append("[КОНТЕНТ] У стовпчику UA не виявлено придатного текстового опису або знайдено технічний HTML/китайський контент.")
                    rows_with_errors.add(i)
                if text_ua.strip() and self.is_suspiciously_short_text(text_ua):
                    row_statuses.append("Підозріло короткий опис")
                    rows_with_short_desc.add(i)

            # Check blacklist terms in UA text
            # Note: Only report first blacklisted term found (intentional - avoids overwhelming with multiple violations)
            text_ua_lower = text_ua.lower()
            for blacklist_entry in blacklist:
                blacklisted_term = blacklist_entry['term']
                exceptions = blacklist_entry.get('exceptions', [])
                action = blacklist_entry.get('action', 'Видалити')
                
                # Check if any exception applies to the whole text
                has_exception = any(exc in text_ua_lower for exc in exceptions if exc)
                if has_exception:
                    continue  # Skip this term entirely if exception found
                
                # Find all occurrences of the term
                idx = 0
                found_in_visible_text = False
                while idx < len(text_ua_lower):
                    idx = text_ua_lower.find(blacklisted_term, idx)
                    if idx == -1:
                        break
                    
                    # Check if this match is in technical HTML context
                    if not self.is_in_technical_context(text_ua, idx):
                        # Match is in visible text - report as error
                        start = max(0, idx - 30)
                        end = min(len(text_ua), idx + len(blacklisted_term) + 30)
                        context = text_ua[start:end].strip()
                        if start > 0:
                            context = "..." + context
                        if end < len(text_ua):
                            context = context + "..."
                        
                        # Different message based on action
                        if action == 'Видалити':
                            row_errors.append(f"[ЗАБОРОНЕНИЙ ТЕРМІН] Знайдено '{blacklisted_term}' у тексті: \"{context}\" (речення буде видалено)")
                        else:  # Підсвітити
                            row_errors.append(f"[ПІДОЗРІЛИЙ ТЕРМІН] Знайдено '{blacklisted_term}' у тексті: \"{context}\" (потребує перевірки)")
                        
                        row_statuses.append("Заборонений термін")
                        rows_with_blacklist.add(i)
                        rows_with_errors.add(i)
                        found_in_visible_text = True
                        break
                    
                    idx += len(blacklisted_term)
                
                if found_in_visible_text:
                    break

            if not html_only_mode:
                for ru_stem, ua_stems in rules.items():
                    match_ru = re.search(r'(?i)\b' + re.escape(ru_stem) + r'[\w-]*\b', text_ru)
                    if match_ru:
                        ru_full_word = match_ru.group(0)
                        for ua_stem in ua_stems:
                            if re.search(r'(?i)\b' + re.escape(ua_stem), text_ua):
                                error_sentence = self.extract_error_sentence(text_ua, ua_stem)
                                is_ignored = False
                                for ig in ignores:
                                    if ig in error_sentence.lower():
                                        is_ignored = True
                                        break
                                if not is_ignored and error_sentence:
                                    row_errors.append(f"[ПОМИЛКА] Значення '{ru_full_word}' хибно перекладено. Речення: \"{error_sentence}\"")
                                    rows_with_errors.add(i)

            df.at[i, errors_col_name] = "\n".join(row_errors)
            df.at[i, checked_ru_col_name] = formatted_ru_text
            df.at[i, checked_ua_col_name] = formatted_ua_text
            
            # Compose status: only include actual error statuses, exclude technical info
            # Priority: real errors first
            if row_statuses:
                df.at[i, status_col_name] = self.compose_status(row_statuses)
            else:
                df.at[i, status_col_name] = "OK"
                rows_ok.add(i)

            if i % 10 == 0:
                self.progress_bar['value'] = i + 1
                self.lbl_progress_status.config(text=f"{os.path.basename(file_path)}: {i + 1}/{total_rows}")

        save_base, ext = os.path.splitext(file_path)
        save_path = f"{save_base}_checked{ext}"

        self.lbl_progress_status.config(text=f"Збереження: {os.path.basename(save_path)}")
        
        # Reorder columns: keep original columns first, then add new columns at the end
        # in order: Помилки, {RU}_checked, {UA}_checked, Статус
        original_cols = [col for col in df.columns if col not in [errors_col_name, status_col_name, checked_ru_col_name, checked_ua_col_name]]
        new_column_order = original_cols + [errors_col_name, checked_ru_col_name, checked_ua_col_name, status_col_name]
        df = df[new_column_order]
        
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Apply formatting to entire table
        from openpyxl.styles import Font, Alignment
        
        # Format header row (row 1)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[1].height = 30
        
        # Format all data rows
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 15
            for cell in ws[row_idx]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Apply highlighting only to _checked columns
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        
        # Find column indices for checked columns
        checked_ru_col_idx = None
        checked_ua_col_idx = None
        for idx, col in enumerate(ws[1], start=1):
            if col.value == checked_ru_col_name:
                checked_ru_col_idx = idx
            elif col.value == checked_ua_col_name:
                checked_ua_col_idx = idx

        for row_idx in rows_with_errors:
            fill_to_use = red_fill if row_idx in rows_with_language_mismatch else yellow_fill
            excel_row_idx = row_idx + 2  # +2 because Excel is 1-based and we have header
            
            # Only highlight the _checked cells
            if checked_ru_col_idx:
                ws.cell(row=excel_row_idx, column=checked_ru_col_idx).fill = fill_to_use
            if checked_ua_col_idx:
                ws.cell(row=excel_row_idx, column=checked_ua_col_idx).fill = fill_to_use

        wb.save(save_path)
        self.progress_bar['value'] = total_rows

        stats = {
            'file_path': file_path,
            'save_path': save_path,
            'total_rows': total_rows,
            'processed_rows': total_rows - len(rows_skipped),
            'rows_with_errors': len(rows_with_errors),
            'need_translation': len(rows_need_translation),
            'technical_html': len(rows_with_technical_html),
            'blacklist': len(rows_with_blacklist),
            'chinese': len(rows_with_chinese),
            'short': len(rows_with_short_desc),
            'skipped': len(rows_skipped),
            'ok_rows': len(rows_ok)
        }
        self.log_message(
            f"Готово: {os.path.basename(file_path)} | рядків: {total_rows}, "
            f"помилки: {stats['rows_with_errors']}, пропущено: {stats['skipped']}, OK: {stats['ok_rows']}"
        )
        return stats

    def run_analysis(self, preferred_ru, preferred_ua):
        try:
            self.log_message("Початок перевірки...")
            self.lbl_progress_status.config(text="Зчитування даних...")

            rules = self.parse_rules()
            ignores = self.parse_ignores()
            blacklist = self.parse_blacklist()
            html_only_mode = self.html_only_var.get()
            skip_processed = self.skip_processed_var.get()
            summaries = []
            result_paths = []

            for file_idx, file_path in enumerate(self.file_paths, start=1):
                self.lbl_progress_status.config(
                    text=f"Файл {file_idx}/{len(self.file_paths)}: {os.path.basename(file_path)}"
                )
                file_stats = self.process_single_file(
                    file_path=file_path,
                    preferred_ru=preferred_ru,
                    preferred_ua=preferred_ua,
                    rules=rules,
                    ignores=ignores,
                    blacklist=blacklist,
                    html_only_mode=html_only_mode,
                    skip_processed=skip_processed
                )
                summaries.append(file_stats)
                result_paths.append(file_stats['save_path'])

            self.last_result_paths = result_paths
            self.root.after(0, lambda: self.btn_open_result.config(state='normal' if self.last_result_paths else 'disabled'))
            summary_text = self.build_summary_text(summaries)
            self.lbl_progress_status.config(text="Готово!")
            self.log_message("Підсумок:\n" + summary_text.replace('\n', ' | '))
            self.root.after(0, lambda: messagebox.showinfo("Підсумок перевірки", summary_text))

        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Помилка", str(e)))
        finally:
            self.root.after(0, lambda: self.set_ui_state(False))

    def start_analysis(self):
        col_ru = self.combo_ru.get()
        col_ua = self.combo_ua.get()
        if not self.file_paths:
            messagebox.showwarning("Увага", "Оберіть хоча б один Excel-файл або папку з файлами.")
            return
        if not col_ru or not col_ua:
            messagebox.showwarning("Увага", "Оберіть стовпчики RU та UA (або дозвольте автопідбір).")
            return
        self.set_ui_state(True)
        self.log_message(
            f"Режим: {'тільки очищення HTML' if self.html_only_var.get() else 'повна перевірка'}; "
            f"пропуск оброблених: {'так' if self.skip_processed_var.get() else 'ні'}."
        )
        threading.Thread(target=self.run_analysis, args=(col_ru, col_ua), daemon=True).start()

    def open_results(self):
        if not self.last_result_paths:
            messagebox.showwarning("Увага", "Ще немає результатів для відкриття.")
            return

        target_path = self.last_result_paths[0]
        if len(self.last_result_paths) > 1:
            target_path = os.path.dirname(target_path)

        try:
            if sys.platform.startswith('win'):
                os.startfile(target_path)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', target_path])
            else:
                subprocess.Popen(['xdg-open', target_path])
            self.log_message(f"Відкрито: {target_path}")
        except Exception as e:
            messagebox.showerror(
                "Помилка",
                f"Не вдалося відкрити результат ({type(e).__name__}):\n{e}\n"
                f"Спробуйте відкрити файл/папку вручну: {target_path}"
            )


if __name__ == "__main__":
    main_root = tk.Tk()
    app = SpellCheckerApp(main_root)
    main_root.mainloop()
