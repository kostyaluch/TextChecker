import os
import re
import sys
import time
import threading
import subprocess
import tkinter as tk
from html import unescape
from tkinter import Toplevel, filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

RULES_FILE = 'translation_rules.txt'
IGNORE_FILE = 'ignore_rules.txt'
BLACKLIST_FILE = 'blacklist_terms.txt'
APP_NAME = "TextChecker"
APP_VERSION = "v15"
MIN_SHORT_DESC_WORDS = 2
MIN_SHORT_DESC_LENGTH = 24
DECIMAL_MAX_INT_DIGITS = 4
DECIMAL_MAX_FRACTION_DIGITS = 2
DECIMAL_CONTEXT_UNITS = r'%|°|кг|г|мг|л|мл|см|мм|м|дюйм|kg|g|mg|l|ml|cm|mm|m|inch|in|x|×'

DEFAULT_RULES = """# Формат: російський_корінь = українські_корені_через_кому
# Наприклад:
питани = харчуванн
лук = цибул
настройк = настройк, настроюванн
вид = вигляд
пользовател = користувач
"""

DEFAULT_IGNORES = """# Додайте сюди фрази (українською), які є ПРАВИЛЬНИМИ винятками.
# Якщо програма знайде цю фразу у реченні, вона проігнорує помилку.
# Кожна фраза з нового рядка:
збалансоване харчування
dитяче харчування
здорове харчування
зелена цибуля
ріпчаста цибуля
""".replace('dитяче', 'дитяче')

DEFAULT_BLACKLIST = """# Додайте сюди слова та фрази (будь-якою мовою), які є ПІДОЗРІЛИМИ або ЗАБОРОНЕНИМИ.
# Якщо програма знайде цю фразу в українському описі, вона позначить рядок як помилковий.
# Також програма автоматично видалить речення або пункт списку, де знайдено заборонений термін.
# Кожна фраза з нового рядка:
www.
http://
@gmail.
e-mail:
копія
репліка
"""

UKRAINIAN_MARKERS = {
    'і', 'та', 'це', 'для', 'що', 'від', 'під', 'при', 'через', 'згідно', 'наявність',
    'зручність', 'технічні', 'обслуговування', 'попередження', 'увага', 'примітка', 'модель'
}
RUSSIAN_MARKERS = {
    'и', 'или', 'это', 'для', 'что', 'при', 'через', 'согласно', 'наличие',
    'удобство', 'технические', 'обслуживание', 'внимание', 'примечание', 'модель'
}
UKRAINIAN_UNIQUE_CHARS = set('іїєґ')
RUSSIAN_UNIQUE_CHARS = set('ыэёъ')
CHINESE_CHAR_PATTERN = re.compile(r'[\u4e00-\u9fff]')
TECHNICAL_HTML_PATTERNS = [
    r'(?is)<style\b[^>]*>.*?</style>',
    r'(?is)<script\b[^>]*>.*?</script>',
    r'(?is)<div\b[^>]*class="[^"]*ssd-module-wrap[^"]*"[^>]*>.*?</div>',
    r'(?is)<div\b[^>]*class=["\']?ssd-module[^>]*>.*?</div>',
    r'(?is)<div\b[^>]*(cssurl|skucode|skudesign|id="zbViewModulesH"|id="zbViewModulesHeight")[^>]*>.*?</div>',
    r'(?is)<input\b[^>]*(zbViewModulesHeight)[^>]*>',
]
HTML_BLACKLIST_EXEMPT_TERMS = ('http://', 'https://', 'ftp://', 'www.')
TECHNICAL_BLACKLIST_CONTEXT_PATTERNS = [
    re.compile(r'(?is)(src|href|poster|data-[\w-]+|cssurl)\s*=\s*["\'][^"\']*$'),
    re.compile(r'(?is)url\(\s*["\']?[^)\"\']*$'),
    re.compile(r'(?is)//[a-z0-9-_.]*$'),
    re.compile(r'(?is)<[^>]*$'),
]
ERROR_STATUS_PREFIXES = ('[МОВА]', '[КОНТЕНТ]', '[ЗАБОРОНЕНИЙ ТЕРМІН]', '[ПОМИЛКА]')


class TextEditor(Toplevel):
    def __init__(self, parent, title, file_path, default_content):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        self.title(title)
        self.geometry("650x450")
        self.file_path = file_path

        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill='both', expand=True)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', side='bottom', pady=(10, 0))

        self.btn_save = ttk.Button(btn_frame, text="Зберегти", command=self.save_and_close)
        self.btn_save.pack(side='right', padx=10)

        self.btn_cancel = ttk.Button(btn_frame, text="Скасувати", command=self.destroy)
        self.btn_cancel.pack(side='right')

        self.txt_content = tk.Text(main_frame, wrap='word', font=('Consolas', 11))
        self.txt_content.pack(fill='both', expand=True, side='top')

        if not os.path.exists(file_path):
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(default_content)

        with open(file_path, 'r', encoding='utf-8') as f:
            self.txt_content.insert('1.0', f.read())

    def save_and_close(self):
        content = self.txt_content.get('1.0', tk.END).strip()
        with open(self.file_path, 'w', encoding='utf-8') as f:
            f.write(content + "\n")
        self.destroy()


class TextCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{APP_NAME} {APP_VERSION}")
        self.root.geometry("980x700")
        self.root.minsize(920, 640)
        self.style = ttk.Style(self.root)
        self.style.theme_use('clam')
        self.configure_styles()

        self.file_paths = []
        self.last_result_paths = []
        self.skip_processed_var = tk.BooleanVar(value=True)
        self.html_only_var = tk.BooleanVar(value=False)
        self.create_widgets()

        if not os.path.exists(RULES_FILE):
            with open(RULES_FILE, 'w', encoding='utf-8') as f:
                f.write(DEFAULT_RULES)
        if not os.path.exists(IGNORE_FILE):
            with open(IGNORE_FILE, 'w', encoding='utf-8') as f:
                f.write(DEFAULT_IGNORES)
        if not os.path.exists(BLACKLIST_FILE):
            with open(BLACKLIST_FILE, 'w', encoding='utf-8') as f:
                f.write(DEFAULT_BLACKLIST)

    def configure_styles(self):
        self.style.configure('Title.TLabel', font=('Segoe UI', 12, 'bold'))
        self.style.configure('TLabelframe', padding=8)
        self.style.configure('TLabelframe.Label', font=('Segoe UI', 10, 'bold'))
        self.style.configure('TButton', padding=(10, 6))
        self.style.configure('TCheckbutton', padding=2)

    def create_widgets(self):
        ttk.Label(self.root, text=f"{APP_NAME} {APP_VERSION}", style='Title.TLabel').pack(
            anchor='w', padx=14, pady=(12, 4)
        )

        file_frame = ttk.LabelFrame(self.root, text="1. Вибір файлів та колонок")
        file_frame.pack(fill='x', padx=10, pady=5)

        file_select_row = ttk.Frame(file_frame)
        file_select_row.pack(fill='x', pady=5)

        self.btn_select_file = ttk.Button(file_select_row, text="📂 Обрати Excel-файли", command=self.select_files)
        self.btn_select_file.pack(side='left', padx=5)

        self.btn_select_folder = ttk.Button(file_select_row, text="📁 Обрати папку", command=self.select_folder)
        self.btn_select_folder.pack(side='left', padx=5)

        self.lbl_file_status = ttk.Label(file_select_row, text="Файли не обрано", foreground="gray")
        self.lbl_file_status.pack(side='left', padx=8, fill='x', expand=True)

        columns_row = ttk.Frame(file_frame)
        columns_row.pack(fill='x', pady=(8, 4))

        ttk.Label(columns_row, text="Стовпчик RU:", font=('Arial', 9, 'bold')).pack(side='left', padx=5)
        self.combo_ru = ttk.Combobox(columns_row, state="disabled", width=25)
        self.combo_ru.pack(side='left', padx=5)

        ttk.Label(columns_row, text="Стовпчик UA:", font=('Arial', 9, 'bold')).pack(side='left', padx=15)
        self.combo_ua = ttk.Combobox(columns_row, state="disabled", width=25)
        self.combo_ua.pack(side='left', padx=5)

        options_frame = ttk.LabelFrame(self.root, text="2. Режими обробки")
        options_frame.pack(fill='x', padx=10, pady=5)

        self.chk_skip_processed = ttk.Checkbutton(
            options_frame,
            text="Пропускати рядки, де _checked вже заповнено",
            variable=self.skip_processed_var
        )
        self.chk_skip_processed.pack(side='left', padx=5)

        self.chk_html_only = ttk.Checkbutton(
            options_frame,
            text="Тільки очищення HTML (без перевірки правил перекладу)",
            variable=self.html_only_var
        )
        self.chk_html_only.pack(side='left', padx=12)

        control_frame = ttk.LabelFrame(self.root, text="3. Керування та словники")
        control_frame.pack(fill='x', padx=10, pady=5)

        self.btn_start_analysis = ttk.Button(control_frame, text="▶ Почати перевірку", command=self.start_analysis, state='disabled')
        self.btn_start_analysis.pack(side='left', padx=5, fill='x', expand=True)

        self.btn_open_result = ttk.Button(control_frame, text="📂 Відкрити результат", command=self.open_results, state='disabled')
        self.btn_open_result.pack(side='left', padx=5)

        self.btn_edit_rules = ttk.Button(control_frame, text="✎ Словник помилок", command=lambda: self.open_editor("Словник помилок", RULES_FILE, DEFAULT_RULES))
        self.btn_edit_rules.pack(side='left', padx=5)

        self.btn_edit_ignores = ttk.Button(control_frame, text="🚫 Словник винятків", command=lambda: self.open_editor("Словник винятків", IGNORE_FILE, DEFAULT_IGNORES))
        self.btn_edit_ignores.pack(side='left', padx=5)

        self.btn_edit_blacklist = ttk.Button(control_frame, text="⛔ Чорний список", command=lambda: self.open_editor("Чорний список термінів", BLACKLIST_FILE, DEFAULT_BLACKLIST))
        self.btn_edit_blacklist.pack(side='left', padx=5)

        progress_frame = ttk.Frame(self.root, padding="5")
        progress_frame.pack(fill='x', padx=10, pady=5)

        self.progress_bar = ttk.Progressbar(progress_frame, orient='horizontal', mode='determinate', length=100)
        self.progress_bar.pack(fill='x', expand=True)

        self.lbl_progress_status = ttk.Label(progress_frame, text="Очікування файлу...", anchor='center')
        self.lbl_progress_status.pack(fill='x', pady=2)

        self.txt_log = tk.Text(self.root, height=10, state='disabled', wrap='word', bg="#f0f0f0", font=('Consolas', 9))
        self.txt_log.pack(fill='both', expand=True, padx=10, pady=10)

        log_scroll = ttk.Scrollbar(self.txt_log, orient='vertical', command=self.txt_log.yview)
        self.txt_log['yscrollcommand'] = log_scroll.set
        log_scroll.pack(side='right', fill='y')

    def log_message(self, message):
        self.txt_log.config(state='normal')
        timestamp = time.strftime("%H:%M:%S")
        self.txt_log.insert(tk.END, f"[{timestamp}] {message}\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state='disabled')

    def open_editor(self, title, file_path, default_content):
        editor = TextEditor(self.root, title, file_path, default_content)
        self.root.wait_window(editor)
        self.log_message(f"Файл '{title}' оновлено.")

    def detect_default_columns(self, columns):
        normalized = {col: str(col).strip().lower() for col in columns}
        ru_priority = ['описание;1', 'описание', 'description;1', 'description ru', 'ru', 'рус']
        ua_priority = ['описание (ua);1', 'опис (ua);1', 'описание ua;1', 'описание (ua)', 'description (ua);1', 'description ua', 'ua', 'uk', 'укр']

        def find_by_priority(priority_list):
            for target in priority_list:
                for original, lowered in normalized.items():
                    if lowered == target:
                        return original
            return ''

        ru_col = find_by_priority(ru_priority)
        ua_col = find_by_priority(ua_priority)

        if not ru_col:
            ru_col = next((
                c for c in columns
                if any(token in normalized[c] for token in ['описание', 'description'])
                and '(ua)' not in normalized[c]
                and 'ua' not in normalized[c]
                and 'uk' not in normalized[c]
            ), '')
        if not ua_col:
            ua_col = next((c for c in columns if any(token in normalized[c] for token in ['ua', 'uk', 'укр', '(ua)'])), '')

        return ru_col, ua_col

    def update_file_selection(self, paths):
        if not paths:
            return
        ordered_paths = list(dict.fromkeys(paths))
        self.file_paths = ordered_paths
        try:
            df_preview = pd.read_excel(ordered_paths[0], nrows=0)
            columns = df_preview.columns.tolist()
            self.combo_ru.config(values=columns, state="readonly")
            self.combo_ua.config(values=columns, state="readonly")
            ru_col, ua_col = self.detect_default_columns(columns)
            if ru_col:
                self.combo_ru.set(ru_col)
            if ua_col:
                self.combo_ua.set(ua_col)
            if len(ordered_paths) == 1:
                self.lbl_file_status.config(text=os.path.basename(ordered_paths[0]), foreground="green")
            else:
                self.lbl_file_status.config(
                    text=f"Обрано файлів: {len(ordered_paths)} (перший: {os.path.basename(ordered_paths[0])})",
                    foreground="green"
                )
            self.log_message(f"Завантажено {len(ordered_paths)} файл(ів). Можна починати перевірку.")
            self.btn_start_analysis.config(state='normal')
        except Exception as e:
            messagebox.showerror("Помилка файлу", f"Не вдалося зчитати файл:\n{e}")

    def select_files(self):
        paths = filedialog.askopenfilenames(
            title="Оберіть один або кілька файлів Excel",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        self.update_file_selection(list(paths))

    def select_folder(self):
        folder = filedialog.askdirectory(title="Оберіть папку з Excel-файлами")
        if not folder:
            return
        paths = []
        for file_name in sorted(os.listdir(folder)):
            if (
                file_name.lower().endswith('.xlsx')
                and not file_name.startswith('~$')
                and not file_name.startswith('.~lock.')
            ):
                paths.append(os.path.join(folder, file_name))
        if not paths:
            messagebox.showwarning("Увага", "У вибраній папці не знайдено .xlsx файлів.")
            return
        self.update_file_selection(paths)

    def set_ui_state(self, is_running):
        state = 'disabled' if is_running else 'normal'
        self.btn_select_file.config(state=state)
        self.btn_select_folder.config(state=state)
        self.btn_edit_rules.config(state=state)
        self.btn_edit_ignores.config(state=state)
        self.btn_edit_blacklist.config(state=state)
        start_state = state if self.file_paths else 'disabled'
        self.btn_start_analysis.config(state=start_state)
        self.chk_skip_processed.config(state=state)
        self.chk_html_only.config(state=state)
        self.combo_ru.config(state="disabled" if is_running else "readonly")
        self.combo_ua.config(state="disabled" if is_running else "readonly")
        if is_running:
            self.btn_open_result.config(state='disabled')

    def parse_rules(self):
        rules = {}
        with open(RULES_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.split('#')[0].strip()
                if '=' in line:
                    ru, ua_list = line.split('=', 1)
                    ua_stems = [ua.strip() for ua in ua_list.split(',')]
                    rules[ru.strip()] = ua_stems
        return rules

    def parse_ignores(self):
        ignores = []
        with open(IGNORE_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.split('#')[0].strip()
                if line:
                    ignores.append(line.lower())
        return ignores

    def parse_blacklist(self):
        blacklist = []
        with open(BLACKLIST_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.split('#')[0].strip()
                if line:
                    blacklist.append(line.lower())
        return blacklist

    def normalize_text(self, text):
        if pd.isna(text):
            return ""
        return self.normalize_common_text_issues(str(text))

    def normalize_common_text_issues(self, text):
        text = str(text).replace('\u00A0', ' ').replace('\u200B', '')
        text = text.replace('“', '"').replace('”', '"').replace('„', '"')
        text = text.replace('’', "'").replace('`', "'")
        text = re.sub(
            rf'(?<!\d)(\d{{1,{DECIMAL_MAX_INT_DIGITS}}}),(\d{{1,{DECIMAL_MAX_FRACTION_DIGITS}}})(?=\s*(?:{DECIMAL_CONTEXT_UNITS}|$|[)\].,;:!?]))',
            r'\1.\2',
            text,
            flags=re.IGNORECASE
        )
        text = re.sub(r'[ \t]+', ' ', text)
        return text.strip()

    def strip_technical_html(self, text):
        text = self.normalize_text(text)
        text = unescape(text)
        text = re.sub(r'\[/?html\]', '', text, flags=re.IGNORECASE)
        for pattern in TECHNICAL_HTML_PATTERNS:
            text = re.sub(pattern, ' ', text)
        text = re.sub(r'(?is)<!--.*?-->', ' ', text)
        return text

    def clean_html_text(self, text):
        text = self.strip_technical_html(text)
        text = re.sub(r'(?is)<style\b[^>]*>.*?</style>', ' ', text)
        text = re.sub(r'(?is)<script\b[^>]*>.*?</script>', ' ', text)
        text = re.sub(r'(?i)(</li>|<br\s*/?>|</p>|</div>)', '. ', text)
        text = re.sub(r'<[^>]+>', ' ', text)
        text = re.sub(r'http[s]?://\S+', ' ', text)
        text = re.sub(r'\b[a-z0-9_-]+\s*\{[^{}]*\}', ' ', text, flags=re.IGNORECASE)
        text = re.sub(r'\s+', ' ', text).strip(' .')
        return text

    def contains_technical_html(self, text):
        raw_text = self.normalize_text(text)
        if not raw_text:
            return False
        lowered = raw_text.lower()
        if '[html]' in lowered or 'ssd-module-wrap' in lowered:
            return True
        return any(re.search(pattern, raw_text) for pattern in TECHNICAL_HTML_PATTERNS)

    def contains_chinese(self, text):
        return bool(CHINESE_CHAR_PATTERN.search(self.normalize_text(text)))

    def is_meaningful_text(self, text):
        cleaned = self.clean_html_text(text)
        if not cleaned:
            return False
        if self.contains_chinese(cleaned):
            return False
        words = re.findall(r'[A-Za-zА-Яа-яІіЇїЄєҐґЁёЪъЫыЭэ0-9]+', cleaned)
        if len(words) <= 1:
            return False
        if len(cleaned) < 12:
            return False
        if 'ssd-module-wrap' in cleaned.lower() or 'background-image' in cleaned.lower():
            return False
        return True

    def is_suspiciously_short_text(self, text):
        cleaned = self.clean_html_text(text)
        if not cleaned:
            return False
        words = re.findall(r'[A-Za-zА-Яа-яІіЇїЄєҐґЁёЪъЫыЭэ]+', cleaned)
        return len(words) <= MIN_SHORT_DESC_WORDS or len(cleaned) < MIN_SHORT_DESC_LENGTH

    def extract_error_sentence(self, text, error_word_stem):
        text_clean = self.clean_html_text(text)
        sentences = re.split(r'(?<=[.!?\n])\s+', text_clean)
        for sentence in sentences:
            if re.search(r'(?i)\b' + re.escape(error_word_stem), sentence):
                return sentence.strip(' .-;:,\t')
        match = re.search(r'(.{0,50}\b' + re.escape(error_word_stem) + r'.{0,50})', text_clean, re.IGNORECASE)
        if match:
            return "..." + match.group(1).strip() + "..."
        return ""

    def sentence_contains_blacklist(self, sentence, blacklist):
        sentence_lower = sentence.lower()
        return any(term in sentence_lower for term in blacklist)

    def is_match_inside_technical_context(self, source_text, match_start, match_end):
        left_context = source_text[max(0, match_start - 120):match_start]
        right_context = source_text[match_end:min(len(source_text), match_end + 120)]
        around_context = source_text[max(0, match_start - 30):min(len(source_text), match_end + 30)]
        combined_context = left_context + source_text[match_start:match_end] + right_context

        for pattern in TECHNICAL_BLACKLIST_CONTEXT_PATTERNS:
            if pattern.search(left_context):
                return True

        if re.search(r'(?is)^\s*["\']?[^"\'>\s]+["\']?\s*>', right_context):
            return True
        if re.search(r'(?is)<img\b', left_context):
            return True
        if re.search(r'(?is)background-image\s*:\s*url\(', left_context):
            return True
        if re.search(r'(?is)cssurl\s*=\s*["\']', left_context):
            return True
        if re.search(r'(?is)src\s*=\s*["\']', left_context):
            return True
        if re.search(r'(?is)href\s*=\s*["\']', left_context):
            return True
        if re.search(r'(?is)<[^>]+>', around_context) and re.search(r'(?is)(src|href|cssurl|style)\s*=', combined_context):
            return True

        return False

    def should_ignore_blacklist_term_in_html(self, source_text, term, match_start, match_end):
        if not source_text or not term:
            return False
        term = term.lower()
        if term in HTML_BLACKLIST_EXEMPT_TERMS or '.' in term:
            return self.is_match_inside_technical_context(source_text, match_start, match_end)
        return False

    def find_blacklist_match(self, source_text, blacklist):
        source_text = self.normalize_text(source_text)
        source_lower = source_text.lower()
        for blacklisted_term in blacklist:
            start_idx = 0
            while True:
                idx = source_lower.find(blacklisted_term, start_idx)
                if idx == -1:
                    break
                match_end = idx + len(blacklisted_term)
                if not self.should_ignore_blacklist_term_in_html(source_text, blacklisted_term, idx, match_end):
                    return blacklisted_term, idx
                start_idx = idx + 1
        return None, -1

    def convert_markdown_bold_to_html(self, text):
        return re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', text)

    def normalize_label(self, label):
        label = re.sub(r'\s+', ' ', label).strip()
        return label[:-1].strip() if label.endswith(':') else label

    def extract_paragraphs(self, source):
        paragraphs = re.findall(r'(?is)<p\b[^>]*>(.*?)</p>', source)
        normalized_paragraphs = []
        for paragraph in paragraphs:
            paragraph = re.sub(r'(?is)<img\b[^>]*>', ' ', paragraph)
            paragraph = self.convert_markdown_bold_to_html(paragraph)
            paragraph = re.sub(r'<br\s*/?>', ' ', paragraph, flags=re.IGNORECASE)
            paragraph = re.sub(r'\s+', ' ', paragraph).strip()
            paragraph_text = self.normalize_common_text_issues(unescape(re.sub(r'<[^>]+>', '', paragraph)).strip())
            if paragraph_text:
                normalized_paragraphs.append((paragraph, paragraph_text))
        if not normalized_paragraphs:
            plain_text = self.clean_html_text(source)
            if plain_text:
                normalized_paragraphs.append((plain_text, plain_text))
        return normalized_paragraphs

    def classify_paragraph(self, paragraph_html, paragraph_text):
        strong_match = re.match(r'^<strong>(.+?)</strong>\s*(.*)$', paragraph_html, flags=re.IGNORECASE | re.DOTALL)
        if strong_match:
            raw_label = strong_match.group(1).strip()
            value = strong_match.group(2).strip()
            label = self.normalize_label(raw_label)
            if value:
                return {'type': 'labeled_value', 'label': label, 'value': value, 'text': paragraph_text}
            return {'type': 'heading', 'label': label, 'text': paragraph_text}

        colon_match = re.match(r'^([^:]{1,80}):\s*(.+)$', paragraph_text)
        if colon_match:
            label = self.normalize_label(colon_match.group(1))
            value = colon_match.group(2).strip()
            if label and value:
                return {'type': 'plain_labeled_value', 'label': label, 'value': value, 'text': paragraph_text}

        return {'type': 'text', 'text': paragraph_text}

    def detect_language(self, text):
        cleaned = self.clean_html_text(text).lower()
        if not cleaned:
            return 'unknown'
        if self.contains_chinese(cleaned):
            return 'zh'

        ua_score = sum(cleaned.count(ch) for ch in UKRAINIAN_UNIQUE_CHARS)
        ru_score = sum(cleaned.count(ch) for ch in RUSSIAN_UNIQUE_CHARS)

        words = re.findall(r'[а-яіїєґёъыэ]+', cleaned, flags=re.IGNORECASE)
        for word in words:
            if word in UKRAINIAN_MARKERS:
                ua_score += 2
            if word in RUSSIAN_MARKERS:
                ru_score += 2

        if ua_score == 0 and ru_score == 0:
            return 'unknown'
        if ua_score > ru_score:
            return 'ua'
        if ru_score > ua_score:
            return 'ru'
        return 'unknown'

    def format_description(self, text, blacklist):
        source = self.strip_technical_html(text)
        paragraphs = self.extract_paragraphs(source)
        if not paragraphs:
            return ""

        result = []
        current_list_items = []
        current_list_signatures = set()

        def flush_list():
            nonlocal current_list_items, current_list_signatures
            if current_list_items:
                result.append('<ul>')
                result.extend(current_list_items)
                result.append('</ul>')
                current_list_items = []
                current_list_signatures = set()

        for paragraph_html, paragraph_text in paragraphs:
            if self.sentence_contains_blacklist(paragraph_text, blacklist):
                continue
            if self.contains_chinese(paragraph_text):
                continue

            item = self.classify_paragraph(paragraph_html, paragraph_text)
            item_type = item['type']

            if item_type == 'heading':
                flush_list()
                result.append(f"<p><strong>{item['label']}:</strong></p>")
                continue

            if item_type in ('labeled_value', 'plain_labeled_value'):
                label = item['label']
                value = self.normalize_common_text_issues(item['value'])

                if self.contains_chinese(label) or self.contains_chinese(value):
                    continue

                if label.lower() in ('увага', 'примітка', 'внимание', 'примечание'):
                    flush_list()
                    result.append(f"<p><strong>{label}:</strong> {value}</p>")
                    continue

                item_signature = (label, value)
                if item_signature not in current_list_signatures:
                    current_list_items.append(f"<li><strong>{label}:</strong> {value}</li>")
                    current_list_signatures.add(item_signature)
                continue

            flush_list()
            result.append(f"<p>{self.normalize_common_text_issues(item['text'])}</p>")

        flush_list()
        formatted = self.cleanup_formatted_html('\n'.join(result).strip())
        return formatted if self.is_meaningful_text(formatted) else ""

    def cleanup_formatted_html(self, html):
        cleaned = re.sub(r'(?is)<p>\s*(?:&nbsp;|\s)*</p>', '', html)
        cleaned = re.sub(r'(?is)<div[^>]*>\s*(?:&nbsp;|\s)*</div>', '', cleaned)
        cleaned = re.sub(r'\n{2,}', '\n', cleaned)
        return cleaned.strip()

    def get_existing_or_create_column(self, df, column_name):
        if column_name not in df.columns:
            df[column_name] = ''
        return column_name

    def resolve_columns_for_file(self, columns, preferred_ru, preferred_ua):
        ru_col = preferred_ru if preferred_ru in columns else ''
        ua_col = preferred_ua if preferred_ua in columns else ''
        if not ru_col or not ua_col:
            detected_ru, detected_ua = self.detect_default_columns(columns)
            ru_col = ru_col or detected_ru
            ua_col = ua_col or detected_ua
        if not ru_col or not ua_col:
            raise ValueError("Не вдалося автоматично визначити колонки RU/UA у файлі.")
        return ru_col, ua_col

    def compose_status_from_errors(self, row_errors):
        statuses = []
        for error in row_errors:
            for prefix in ERROR_STATUS_PREFIXES:
                if error.startswith(prefix):
                    statuses.append(prefix.strip('[]'))
                    break
        ordered = []
        for status in statuses:
            if status not in ordered:
                ordered.append(status)
        return " | ".join(ordered) if ordered else "OK"

    def build_summary_text(self, summaries):
        total_files = len(summaries)
        totals = {
            'total_rows': 0,
            'processed_rows': 0,
            'rows_with_errors': 0,
            'need_translation': 0,
            'technical_html': 0,
            'blacklist': 0,
            'chinese': 0,
            'short': 0,
            'skipped': 0,
            'ok_rows': 0
        }
        for stats in summaries:
            for key in totals:
                totals[key] += stats.get(key, 0)
        return (
            f"Оброблено файлів: {total_files}\n"
            f"Усього рядків: {totals['total_rows']}\n"
            f"Опрацьовано рядків: {totals['processed_rows']}\n"
            f"Рядків з помилками: {totals['rows_with_errors']}\n"
            f"Потрібен переклад: {totals['need_translation']}\n"
            f"Технічний HTML: {totals['technical_html']}\n"
            f"Заборонений термін: {totals['blacklist']}\n"
            f"Китайський текст: {totals['chinese']}\n"
            f"Підозріло короткий опис: {totals['short']}\n"
            f"Пропущено (вже оброблено): {totals['skipped']}\n"
            f"OK: {totals['ok_rows']}"
        )

    def reorder_result_columns(self, df, errors_col_name, checked_ru_col_name, checked_ua_col_name, status_col_name):
        rename_map = {errors_col_name: 'Помилки'}
        df = df.rename(columns=rename_map)
        preferred_order = ['Помилки', checked_ru_col_name, checked_ua_col_name, status_col_name]
        existing_preferred = [column for column in preferred_order if column in df.columns]
        remaining_columns = [column for column in df.columns if column not in existing_preferred]
        return df[remaining_columns + existing_preferred]

    def apply_worksheet_formatting(self, ws, checked_ru_excel_col, checked_ua_excel_col, rows_with_language_mismatch, rows_with_other_errors):
        header_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            ws.row_dimensions[row_idx].height = 30 if row_idx == 1 else 15
            for cell in row:
                cell.alignment = wrap_alignment
                if row_idx == 1:
                    cell.font = header_font

        for row_idx in rows_with_other_errors:
            for column_idx in (checked_ru_excel_col, checked_ua_excel_col):
                ws.cell(row=row_idx + 2, column=column_idx).fill = yellow_fill

        for row_idx in rows_with_language_mismatch:
            for column_idx in (checked_ru_excel_col, checked_ua_excel_col):
                ws.cell(row=row_idx + 2, column=column_idx).fill = red_fill

    def process_single_file(self, file_path, preferred_ru, preferred_ua, rules, ignores, blacklist, html_only_mode, skip_processed):
        self.log_message(f"Обробка файлу: {os.path.basename(file_path)}")
        df = pd.read_excel(file_path)
        total_rows = len(df)
        self.progress_bar.config(maximum=max(total_rows, 1), value=0)

        col_ru, col_ua = self.resolve_columns_for_file(df.columns, preferred_ru, preferred_ua)
        errors_col_name = self.get_existing_or_create_column(df, 'Помилки перекладу')
        status_col_name = self.get_existing_or_create_column(df, 'Статус')
        checked_ru_col_name = self.get_existing_or_create_column(df, f'{col_ru}_checked')
        checked_ua_col_name = self.get_existing_or_create_column(df, f'{col_ua}_checked')

        rows_with_errors = set()
        rows_with_language_mismatch = set()
        rows_need_translation = set()
        rows_with_technical_html = set()
        rows_with_blacklist = set()
        rows_with_chinese = set()
        rows_with_short_desc = set()
        rows_skipped = set()
        rows_ok = set()
        rows_with_other_errors = set()

        for i, row in df.iterrows():
            text_ru = self.normalize_text(row.get(col_ru, ''))
            text_ua = self.normalize_text(row.get(col_ua, ''))
            row_errors = []

            existing_ru_checked = self.normalize_text(row.get(checked_ru_col_name, ''))
            existing_ua_checked = self.normalize_text(row.get(checked_ua_col_name, ''))
            if skip_processed and (existing_ru_checked.strip() or existing_ua_checked.strip()):
                rows_skipped.add(i)
                df.at[i, status_col_name] = "Пропущено (вже оброблено)"
                continue

            ru_detected_lang = self.detect_language(text_ru)
            ua_detected_lang = self.detect_language(text_ua)
            ru_has_technical = self.contains_technical_html(text_ru)
            ua_has_technical = self.contains_technical_html(text_ua)
            ru_has_chinese = self.contains_chinese(text_ru)
            ua_has_chinese = self.contains_chinese(text_ua)

            if ru_has_technical or ua_has_technical:
                rows_with_technical_html.add(i)
            if ru_has_chinese or ua_has_chinese:
                rows_with_chinese.add(i)

            if text_ru.strip() and ru_detected_lang in ('ua', 'zh') and not html_only_mode:
                if ru_detected_lang == 'zh':
                    row_errors.append("[МОВА] У стовпчику RU виявлено китайський текст. Потрібен переклад російською.")
                else:
                    row_errors.append("[МОВА] У стовпчику RU виявлено український текст. Потрібен переклад російською.")
                formatted_ru_text = ""
                rows_need_translation.add(i)
                rows_with_errors.add(i)
                rows_with_language_mismatch.add(i)
            else:
                formatted_ru_text = self.format_description(text_ru, blacklist)
                if text_ru.strip() and not formatted_ru_text:
                    row_errors.append("[КОНТЕНТ] У стовпчику RU не виявлено придатного текстового опису або знайдено технічний HTML/китайський контент.")
                    rows_with_errors.add(i)
                    rows_with_other_errors.add(i)
                if text_ru.strip() and self.is_suspiciously_short_text(text_ru):
                    rows_with_short_desc.add(i)

            if text_ua.strip() and ua_detected_lang in ('ru', 'zh') and not html_only_mode:
                if ua_detected_lang == 'zh':
                    row_errors.append("[МОВА] У стовпчику UA виявлено китайський текст. Потрібен переклад українською.")
                else:
                    row_errors.append("[МОВА] У стовпчику UA виявлено російський текст. Потрібен переклад українською.")
                formatted_ua_text = ""
                rows_need_translation.add(i)
                rows_with_errors.add(i)
                rows_with_language_mismatch.add(i)
            else:
                formatted_ua_text = self.format_description(text_ua, blacklist)
                if text_ua.strip() and not formatted_ua_text:
                    row_errors.append("[КОНТЕНТ] У стовпчику UA не виявлено придатного текстового опису або знайдено технічний HTML/китайський контент.")
                    rows_with_errors.add(i)
                    rows_with_other_errors.add(i)
                if text_ua.strip() and self.is_suspiciously_short_text(text_ua):
                    rows_with_short_desc.add(i)

            matched_blacklist_term, matched_idx = self.find_blacklist_match(text_ua, blacklist)
            if matched_blacklist_term:
                start = max(0, matched_idx - 30)
                end = min(len(text_ua), matched_idx + len(matched_blacklist_term) + 30)
                context = text_ua[start:end].strip()
                if start > 0:
                    context = "..." + context
                if end < len(text_ua):
                    context = context + "..."
                row_errors.append(f"[ЗАБОРОНЕНИЙ ТЕРМІН] Знайдено '{matched_blacklist_term}' у тексті: \"{context}\"")
                rows_with_blacklist.add(i)
                rows_with_errors.add(i)
                rows_with_other_errors.add(i)

            if not html_only_mode:
                for ru_stem, ua_stems in rules.items():
                    match_ru = re.search(r'(?i)\b' + re.escape(ru_stem) + r'[\w-]*\b', text_ru)
                    if match_ru:
                        ru_full_word = match_ru.group(0)
                        for ua_stem in ua_stems:
                            if re.search(r'(?i)\b' + re.escape(ua_stem), text_ua):
                                error_sentence = self.extract_error_sentence(text_ua, ua_stem)
                                is_ignored = False
                                for ig in ignores:
                                    if ig in error_sentence.lower():
                                        is_ignored = True
                                        break
                                if not is_ignored and error_sentence:
                                    row_errors.append(f"[ПОМИЛКА] Значення '{ru_full_word}' хибно перекладено. Речення: \"{error_sentence}\"")
                                    rows_with_errors.add(i)
                                    rows_with_other_errors.add(i)

            df.at[i, errors_col_name] = "\n".join(row_errors)
            df.at[i, checked_ru_col_name] = formatted_ru_text
            df.at[i, checked_ua_col_name] = formatted_ua_text
            df.at[i, status_col_name] = self.compose_status_from_errors(row_errors)

            if row_errors:
                rows_with_errors.add(i)
            else:
                rows_ok.add(i)

            if i % 10 == 0:
                self.progress_bar['value'] = i + 1
                self.lbl_progress_status.config(text=f"{os.path.basename(file_path)}: {i + 1}/{total_rows}")

        df = self.reorder_result_columns(df, errors_col_name, checked_ru_col_name, checked_ua_col_name, status_col_name)

        save_base, ext = os.path.splitext(file_path)
        save_path = f"{save_base}_checked{ext}"

        self.lbl_progress_status.config(text=f"Збереження: {os.path.basename(save_path)}")
        wb = Workbook()
        ws = wb.active
        for row_data in dataframe_to_rows(df, index=False, header=True):
            ws.append(row_data)

        checked_ru_excel_col = list(df.columns).index(checked_ru_col_name) + 1
        checked_ua_excel_col = list(df.columns).index(checked_ua_col_name) + 1
        self.apply_worksheet_formatting(ws, checked_ru_excel_col, checked_ua_excel_col, rows_with_language_mismatch, rows_with_other_errors)

        wb.save(save_path)
        self.progress_bar['value'] = total_rows

        stats = {
            'file_path': file_path,
            'save_path': save_path,
            'total_rows': total_rows,
            'processed_rows': total_rows - len(rows_skipped),
            'rows_with_errors': len(rows_with_errors),
            'need_translation': len(rows_need_translation),
            'technical_html': len(rows_with_technical_html),
            'blacklist': len(rows_with_blacklist),
            'chinese': len(rows_with_chinese),
            'short': len(rows_with_short_desc),
            'skipped': len(rows_skipped),
            'ok_rows': len(rows_ok)
        }
        self.log_message(
            f"Готово: {os.path.basename(file_path)} | рядків: {total_rows}, "
            f"помилки: {stats['rows_with_errors']}, пропущено: {stats['skipped']}, OK: {stats['ok_rows']}"
        )
        return stats

    def run_analysis(self, preferred_ru, preferred_ua):
        try:
            self.log_message("Початок перевірки...")
            self.lbl_progress_status.config(text="Зчитування даних...")

            rules = self.parse_rules()
            ignores = self.parse_ignores()
            blacklist = self.parse_blacklist()
            html_only_mode = self.html_only_var.get()
            skip_processed = self.skip_processed_var.get()
            summaries = []
            result_paths = []

            for file_idx, file_path in enumerate(self.file_paths, start=1):
                self.lbl_progress_status.config(
                    text=f"Файл {file_idx}/{len(self.file_paths)}: {os.path.basename(file_path)}"
                )
                file_stats = self.process_single_file(
                    file_path=file_path,
                    preferred_ru=preferred_ru,
                    preferred_ua=preferred_ua,
                    rules=rules,
                    ignores=ignores,
                    blacklist=blacklist,
                    html_only_mode=html_only_mode,
                    skip_processed=skip_processed
                )
                summaries.append(file_stats)
                result_paths.append(file_stats['save_path'])

            self.last_result_paths = result_paths
            self.root.after(0, lambda: self.btn_open_result.config(state='normal' if self.last_result_paths else 'disabled'))
            summary_text = self.build_summary_text(summaries)
            self.lbl_progress_status.config(text="Готово!")
            self.log_message("Підсумок:\n" + summary_text.replace('\n', ' | '))
            self.root.after(0, lambda: messagebox.showinfo(f"Підсумок {APP_NAME}", summary_text))

        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror(APP_NAME, str(e)))
        finally:
            self.root.after(0, lambda: self.set_ui_state(False))

    def start_analysis(self):
        col_ru = self.combo_ru.get()
        col_ua = self.combo_ua.get()
        if not self.file_paths:
            messagebox.showwarning("Увага", "Оберіть хоча б один Excel-файл або папку з файлами.")
            return
        if not col_ru or not col_ua:
            messagebox.showwarning("Увага", "Оберіть стовпчики RU та UA (або дозвольте автопідбір).")
            return
        self.set_ui_state(True)
        self.log_message(
            f"Режим: {'тільки очищення HTML' if self.html_only_var.get() else 'повна перевірка'}; "
            f"пропуск оброблених: {'так' if self.skip_processed_var.get() else 'ні'}."
        )
        threading.Thread(target=self.run_analysis, args=(col_ru, col_ua), daemon=True).start()

    def open_results(self):
        if not self.last_result_paths:
            messagebox.showwarning("Увага", "Ще немає результатів для відкриття.")
            return

        target_path = self.last_result_paths[0]
        if len(self.last_result_paths) > 1:
            target_path = os.path.dirname(target_path)

        try:
            if sys.platform.startswith('win'):
                os.startfile(target_path)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', target_path])
            else:
                subprocess.Popen(['xdg-open', target_path])
            self.log_message(f"Відкрито: {target_path}")
        except Exception as e:
            messagebox.showerror(
                "Помилка",
                f"Не вдалося відкрити результат ({type(e).__name__}):\n{e}\n"
                f"Спробуйте відкрити файл/папку вручну: {target_path}"
            )


if __name__ == "__main__":
    main_root = tk.Tk()
    app = TextCheckerApp(main_root)
    main_root.mainloop()
